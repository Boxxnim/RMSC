"""
Cohort Tracking Module for Machine Perfusion Systematic Review
Manages study registry, duplicate detection, and outcome matrix generation.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class CohortTracker:
    """Manages cohort_tracking.xlsx for systematic review"""
    
    # Sheet names
    STUDY_REGISTRY = "Study Registry"
    COHORT_LINKAGE = "Cohort Linkage"
    OUTCOME_MATRIX = "Outcome Matrix"
    CHANGE_LOG = "Change Log"
    
    # Outcome fields to track (must match Excel header order)
    OUTCOMES = [
        "EAD", "NAS", "TBC", "ACR", "PNF", "HAT", 
        "Retx", "RRT", "1yr Graft", "1yr Patient",
        "AKI", "PRS", "Major Comp", "Hospital Stay", "ICU Stay"
    ]
    
    def __init__(self, tracking_file: str):
        """Load existing cohort_tracking.xlsx"""
        self.tracking_file = Path(tracking_file)
        
        if self.tracking_file.exists():
            self.wb = openpyxl.load_workbook(tracking_file)
        else:
            raise FileNotFoundError(f"Tracking file not found: {tracking_file}")
    
    def study_exists(self, study_id: str) -> bool:
        """Check if study already exists in registry"""
        ws = self.wb[self.STUDY_REGISTRY]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == study_id:
                return True
        return False
    
    def check_duplicates(self, study_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Check for potential duplicate studies based on NCT ID
        
        Returns list of potential matches with overlap info.
        Primary method: NCT ID matching
        """
        potential_duplicates = []
        registry_id = study_info.get("registry_id", "")
        
        if not registry_id:
            return []
        
        ws = self.wb[self.STUDY_REGISTRY]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        
        nct_col = headers.index("Registry ID (NCT)") + 1 if "Registry ID (NCT)" in headers else None
        study_id_col = 1
        
        if nct_col:
            for row in range(2, ws.max_row + 1):
                existing_nct = ws.cell(row=row, column=nct_col).value
                existing_study_id = ws.cell(row=row, column=study_id_col).value
                
                if existing_nct and registry_id.upper() in str(existing_nct).upper():
                    potential_duplicates.append({
                        "existing_study_id": existing_study_id,
                        "match_type": "NCT Match",
                        "existing_nct": existing_nct,
                        "new_nct": registry_id
                    })
        
        return potential_duplicates
    
    def register_study(self, extraction_result: Dict[str, Any]) -> Dict[str, Any]:
        """Register a study in Study Registry sheet
        
        Returns registration result with duplicate warnings if any.
        """
        full_result = extraction_result.get("full_extraction", extraction_result)
        study_chars = full_result.get("study_characteristics", {})
        
        study_id = study_chars.get("study_id", "Unknown")
        
        # Check if already exists
        if self.study_exists(study_id):
            return {
                "status": "exists",
                "study_id": study_id,
                "message": f"Study {study_id} already exists in registry"
            }
        
        # Check for duplicates
        duplicates = self.check_duplicates(study_chars)
        
        # Add to registry
        ws = self.wb[self.STUDY_REGISTRY]
        new_row = ws.max_row + 1
        
        # Map extraction result to registry columns
        ws.cell(row=new_row, column=1, value=study_id)
        ws.cell(row=new_row, column=2, value=study_chars.get("first_author"))
        ws.cell(row=new_row, column=3, value=study_chars.get("year"))
        ws.cell(row=new_row, column=4, value=study_chars.get("title"))
        ws.cell(row=new_row, column=5, value=study_chars.get("journal"))
        ws.cell(row=new_row, column=6, value=study_chars.get("doi"))
        ws.cell(row=new_row, column=7, value=study_chars.get("study_design"))
        ws.cell(row=new_row, column=8, value=", ".join(study_chars.get("centers", []) or []))
        ws.cell(row=new_row, column=9, value=", ".join(study_chars.get("countries", []) or []))
        
        # Enrollment period
        start = study_chars.get("enrollment_period_start", "")
        end = study_chars.get("enrollment_period_end", "")
        period = f"{start}/{end}" if start or end else ""
        ws.cell(row=new_row, column=10, value=period)
        
        ws.cell(row=new_row, column=11, value=study_chars.get("registry_id"))
        ws.cell(row=new_row, column=12, value="-")  # Parent Study ID
        ws.cell(row=new_row, column=13, value="Primary")  # Relationship Type
        ws.cell(row=new_row, column=14, value=study_chars.get("n_total"))
        ws.cell(row=new_row, column=15, value=study_chars.get("n_intervention"))
        ws.cell(row=new_row, column=16, value=study_chars.get("n_control"))
        ws.cell(row=new_row, column=17, value=study_chars.get("intervention_type"))
        
        # Notes with duplicate warnings
        notes = []
        if duplicates:
            for dup in duplicates:
                notes.append(f"⚠️ Potential duplicate: {dup['existing_study_id']} ({dup['match_type']})")
        ws.cell(row=new_row, column=18, value="; ".join(notes) if notes else "")
        
        return {
            "status": "registered",
            "study_id": study_id,
            "row": new_row,
            "duplicates": duplicates
        }
    
    def generate_outcome_matrix_row(self, extraction_result: Dict[str, Any]) -> Dict[str, str]:
        """Generate Outcome Matrix row from extraction result
        
        reported: true → "Include"
        reported: false → "N/A"
        """
        full_result = extraction_result.get("full_extraction", extraction_result)
        study_id = full_result.get("study_characteristics", {}).get("study_id", "Unknown")
        outcome_data = full_result.get("outcome_data", {})
        
        # Map schema field names to matrix column names
        field_mapping = {
            "ead": "EAD",
            "nas": "NAS", 
            "tbc": "TBC",
            "pnf": "PNF",
            "acr": "ACR",
            "hat": "HAT",
            "retransplantation": "Retx",
            "rrt": "RRT",
            "aki": "AKI",
            "prs": "PRS",
            "major_complications": "Major Comp",
            "graft_survival_1yr": "1yr Graft",
            "patient_survival_1yr": "1yr Patient",
            "hospital_stay_days": "Hospital Stay",
            "icu_stay_days": "ICU Stay",
        }
        
        row_data = {"Study_ID": study_id}
        
        for field, col_name in field_mapping.items():
            outcome = outcome_data.get(field, {})
            if isinstance(outcome, dict):
                reported = outcome.get("reported", False)
                row_data[col_name] = "Include" if reported else "N/A"
            else:
                row_data[col_name] = "N/A"
        
        return row_data
    
    def update_outcome_matrix(self, extraction_result: Dict[str, Any]) -> Dict[str, Any]:
        """Add or update row in Outcome Matrix sheet"""
        row_data = self.generate_outcome_matrix_row(extraction_result)
        study_id = row_data["Study_ID"]
        
        ws = self.wb[self.OUTCOME_MATRIX]
        
        # Check if row exists
        existing_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == study_id:
                existing_row = row
                break
        
        target_row = existing_row if existing_row else ws.max_row + 1
        
        # Write row
        ws.cell(row=target_row, column=1, value=study_id)
        for i, outcome in enumerate(self.OUTCOMES, start=2):
            value = row_data.get(outcome, "N/A")
            ws.cell(row=target_row, column=i, value=value)
        
        return {
            "status": "updated" if existing_row else "added",
            "study_id": study_id,
            "row": target_row,
            "data": row_data
        }
    
    def add_change_log(self, study_ids: List[str], action: str, rationale: str, author: str = "LLM"):
        """Add entry to Change Log"""
        ws = self.wb[self.CHANGE_LOG]
        new_row = ws.max_row + 1
        
        ws.cell(row=new_row, column=1, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=new_row, column=2, value=author)
        ws.cell(row=new_row, column=3, value=", ".join(study_ids))
        ws.cell(row=new_row, column=4, value=action)
        ws.cell(row=new_row, column=5, value=rationale)
        ws.cell(row=new_row, column=6, value="Pending")
    
    def save(self):
        """Save changes to file"""
        self.wb.save(self.tracking_file)
    
    def process_extraction(self, extraction_result: Dict[str, Any]) -> Dict[str, Any]:
        """Process extraction result: register study and update outcome matrix"""
        
        # Register in Study Registry
        registry_result = self.register_study(extraction_result)
        
        # Update Outcome Matrix
        matrix_result = self.update_outcome_matrix(extraction_result)
        
        # Add to change log
        study_id = registry_result.get("study_id", "Unknown")
        if registry_result["status"] == "registered":
            self.add_change_log(
                [study_id],
                "Auto-registered from LLM extraction",
                f"Extracted with extractor_gemini.py"
            )
        
        # Save
        self.save()
        
        return {
            "registry": registry_result,
            "outcome_matrix": matrix_result,
            "duplicates": registry_result.get("duplicates", [])
        }
