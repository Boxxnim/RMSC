#!/usr/bin/env python3
"""
Data Extractor (Gemini Version) for Machine Perfusion Systematic Review
목적: 메타분석용 데이터 수집 (Gemini API 사용)
최적화: Quick Screen 제거 → Full Extraction + RoB Extraction (2 API calls)
"""

import json
import os
import re
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List

# Load .env file if exists
from dotenv import load_dotenv
load_dotenv()

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Local imports
from prompts_v2_data_collection import (
    get_data_extraction_prompt,
    get_quick_screen_prompt,
    get_rob_extraction_prompt,
)
from pdf_utils import extract_text_from_pdf, extract_text_with_ocr_fallback
from schemas_gemini import FULL_EXTRACTION_SCHEMA, ROB2_EXTRACTION_SCHEMA, ROBINS_I_EXTRACTION_SCHEMA
from tool_definitions import INITIAL_TOOLS, get_extraction_tools
from skills import get_skill
from extraction_state import ExtractionState


# =============================================================================
# Configuration
# =============================================================================

DEFAULT_MODEL = "gemini-3-flash-preview"
TEMPLATE_PATH = "data_extraction_template_v2.xlsx"

# Sheet names in template
SHEETS = {
    "characteristics": "Study_Characteristics",
    "perfusion": "Perfusion_Settings",
    "time": "Time_Metrics",
    "outcomes": "Outcome_Data",
    "continuous": "Continuous_Outcomes",
    "rob2": "RoB2_RCT",
    "robins": "ROBINS_I_NRS",
    "notes": "Extraction_Notes"
}


# =============================================================================
# Gemini LLM Client
# =============================================================================

class GeminiClient:
    def __init__(self, model: str = DEFAULT_MODEL):
        # API key from environment
        api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GOOGLE_API_KEY or GEMINI_API_KEY environment variable not set")
        
        # New google.genai SDK - client-based approach
        self.client = genai.Client(api_key=api_key)
        self.model_name = model
    
    def extract(self, prompt: str, schema: types.Schema = None) -> Dict[str, Any]:
        """Call Gemini API with structured output schema (text-only)
        
        Args:
            prompt: The prompt to send
            schema: types.Schema for structured output (enforced JSON format)
        """
        
        # Gemini 3 Flash Preview config:
        # - temperature=0 for deterministic extraction
        # - thinking_level="high" for thorough analysis
        # - response_schema for structured JSON output (no max_output_tokens limit)
        config = types.GenerateContentConfig(
            temperature=0,  # Deterministic output
            thinking_config=types.ThinkingConfig(
                thinking_level="medium"  # Medium to prevent thinking loops
            ),
            response_mime_type="application/json",
            response_schema=schema if schema else None,
        )
        
        try:
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=prompt,
                config=config
            )
            
            # With response_schema, response should be valid JSON matching schema
            text = response.text
            
            try:
                return json.loads(text)
            except json.JSONDecodeError as e:
                # Fallback: try to extract JSON from response
                json_match = re.search(r'\{[\s\S]*\}', text)
                if json_match:
                    try:
                        return json.loads(json_match.group())
                    except json.JSONDecodeError:
                        pass
                print(f"JSON parse error: {e}")
                return {"raw_response": text, "parse_error": str(e)}
                
        except Exception as e:
            print(f"API call error: {e}")
            # Fallback without thinking_config if not supported
            if "thinking" in str(e).lower():
                print("  Retrying without thinking config...")
                fallback_config = types.GenerateContentConfig(
                    temperature=0,
                    response_mime_type="application/json",
                    response_schema=schema if schema else None,
                )
                response = self.client.models.generate_content(
                    model=self.model_name,
                    contents=prompt,
                    config=fallback_config
                )
                text = response.text
                try:
                    return json.loads(text)
                except json.JSONDecodeError as e2:
                    json_match = re.search(r'\{[\s\S]*\}', text)
                    if json_match:
                        try:
                            return json.loads(json_match.group())
                        except json.JSONDecodeError:
                            pass
                    return {"raw_response": text, "parse_error": str(e2)}
            raise
    
    def extract_from_pdf(self, pdf_paths: List[str], prompt: str, schema: types.Schema = None) -> Dict[str, Any]:
        """Call Gemini API with PDF files directly (multimodal)
        
        Args:
            pdf_paths: List of paths to PDF files
            prompt: The prompt to send
            schema: types.Schema for structured output
        """
        
        # Build contents with PDF parts + prompt
        contents = []
        
        for pdf_path in pdf_paths:
            with open(pdf_path, 'rb') as f:
                pdf_data = f.read()
            contents.append(
                types.Part.from_bytes(
                    data=pdf_data,
                    mime_type='application/pdf'
                )
            )
        
        # Add prompt at the end
        contents.append(prompt)
        
        config = types.GenerateContentConfig(
            temperature=0,
            thinking_config=types.ThinkingConfig(
                thinking_level="medium"
            ),
            response_mime_type="application/json",
            response_schema=schema if schema else None,
        )
        
        try:
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=contents,
                config=config
            )
            
            text = response.text
            
            try:
                return json.loads(text)
            except json.JSONDecodeError as e:
                json_match = re.search(r'\{[\s\S]*\}', text)
                if json_match:
                    try:
                        return json.loads(json_match.group())
                    except json.JSONDecodeError:
                        pass
                print(f"JSON parse error: {e}")
                return {"raw_response": text, "parse_error": str(e)}
                
        except Exception as e:
            print(f"API call error: {e}")
            if "thinking" in str(e).lower():
                print("  Retrying without thinking config...")
                fallback_config = types.GenerateContentConfig(
                    temperature=0,
                    response_mime_type="application/json",
                    response_schema=schema if schema else None,
                )
                response = self.client.models.generate_content(
                    model=self.model_name,
                    contents=contents,
                    config=fallback_config
                )
                text = response.text
                try:
                    return json.loads(text)
                except json.JSONDecodeError as e2:
                    json_match = re.search(r'\{[\s\S]*\}', text)
                    if json_match:
                        try:
                            return json.loads(json_match.group())
                        except json.JSONDecodeError:
                            pass
                    return {"raw_response": text, "parse_error": str(e2)}
            raise
    
    def extract_with_tools(self, pdf_paths: List[str], rob_files: List[str] = None,
                           verbose: bool = False, 
                           save_state: bool = False, state_path: Optional[str] = None) -> Tuple[Dict[str, Any], Optional[ExtractionState]]:
        """Single API call extraction with tool calling and JIT skill injection.
        
        Workflow:
        1. Model reads PDF and analyzes study design
        2. Model calls get_rob_skill() to get appropriate RoB instructions
        3. We inject the skill AND RoB-related files (disclosure, protocol, COI)
        4. Model calls submit_extraction() with complete data
        
        Args:
            pdf_paths: List of PDF files to process
            rob_files: List of RoB-related PDFs (disclosure, protocol, COI) to inject with skill
            verbose: If True, print thinking process and tool calls
            save_state: If True, save extraction state for later interactive review
            state_path: Optional path for state file (auto-generated if not provided)
            
        Returns:
            Tuple of (extraction data dict, ExtractionState or None if save_state=False)
        """
        
        # Build contents with PDF parts
        contents = []
        for pdf_path in pdf_paths:
            with open(pdf_path, 'rb') as f:
                pdf_data = f.read()
            contents.append(
                types.Part.from_bytes(
                    data=pdf_data,
                    mime_type='application/pdf'
                )
            )
        
        # System prompt for tool-based extraction
        system_prompt = """You are a systematic review data extractor for liver transplantation machine perfusion studies.

## YOUR TASK
Extract ALL available data from this paper for meta-analysis and assess Risk of Bias.

## WORKFLOW
1. Read the paper carefully and determine the study design (RCT or NRS)
2. Call get_rob_skill(study_type, study_id) to get detailed RoB assessment instructions
3. Extract all data following the returned instructions
4. Call submit_extraction() with complete results

## STUDY CONTEXT
This systematic review compares:
- HOPE (Hypothermic Oxygenated Perfusion) vs SCS (Static Cold Storage)
- NMP (Normothermic Machine Perfusion) vs SCS
- In extended criteria donor (ECD) liver transplantation

## OUTCOME DEFINITIONS AND HETEROGENEITY
Different papers may use different definitions for the same outcome. Handle this as follows:
1. ALWAYS record the paper's EXACT definition in the 'definition' field
2. If the paper's definition differs from standard criteria, still extract the data
3. For composite outcomes (like TBC/Total Biliary Complications), sum ALL component types mentioned
4. When multiple time points are available, prefer 6 months for NAS, 1 year for survival
5. If definition is unclear or non-standard, note this in 'extraction_notes.unclear_items_for_review'

Common definition variations to watch for:
- EAD: Olthoff criteria vs other criteria (always specify which)
- NAS: different follow-up periods (6mo, 12mo, etc.)
- PRS: 20% vs 30% MAP threshold
- AKI: KDIGO vs RIFLE vs AKIN criteria
- Major complications: Clavien-Dindo grade threshold

## NRS: MATCHED DATA ONLY
For non-randomized studies (NRS):
1. Extract data from MATCHED cohorts ONLY, not unmatched/entire cohort
2. If the paper reports both matched and unmatched results, use ONLY the matched results
3. Record matching details in study_characteristics:
   - matching_method: PSM, IPTW, case-control, caliper matching, etc.
   - matching_ratio: 1:1, 1:2, 1:3, etc.
   - matching_variables: list of variables used (age, MELD, DRI, donor type, etc.)
4. n_intervention and n_control should reflect the MATCHED sample sizes
5. If no matching was performed, note this in extraction_notes

## ECD/DCD SUBGROUP PRIORITY
This systematic review focuses on EXTENDED CRITERIA DONORS (ECD). 

### ECD Definition:
Extended Criteria Donors include:
- Age ≥60 years, OR
- Age 50-60 with 2+ of: hypertension, terminal creatinine >1.5, CVA as cause of death
- DCD (Donation after Circulatory Death) donors
- High DRI (Donor Risk Index) donors (typically DRI >1.5 or >1.7)

### Extraction Rules:
1. If the paper includes both standard and ECD/DCD donors:
   - PRIORITIZE the DCD or ECD SUBGROUP data over the full cohort
   - If DCD-specific outcomes are reported separately, use those numbers
2. If the paper reports results for DCD donors separately (e.g., "DCD subgroup analysis"):
   - Extract from the DCD subgroup, NOT the entire cohort
3. Record in study_characteristics.donor_type: "DCD", "DBD", "ECD-DBD", or "Mixed"
4. If only pooled data is available (no ECD/DCD breakdown), extract the pooled data 
   and note "No DCD/ECD subgroup available - using pooled data" in extraction_notes.general_notes

## UNCERTAINTY FLAGGING
When you encounter ambiguity, FLAG it explicitly:
1. If MULTIPLE COHORTS exist (e.g., matched vs unmatched, different time periods):
   - Add to unclear_items_for_review: "Multiple cohorts: [describe options and which you chose]"
2. If N values seem inconsistent across outcomes:
   - Add to data_quality_concerns: "N inconsistency: [describe]"
3. If outcome definition is unclear:
   - Add to unclear_items_for_review: "[Outcome] definition unclear: [describe]"

## N VALUE CONSISTENCY
- n_intervention and n_control should be CONSISTENT across all outcomes
- If you extract EAD with N=50/25, all other outcomes should use the same N
- If N differs for survival outcomes (due to follow-up), note this in general_notes

## COHORT IDENTIFICATION (CRITICAL FOR OVERLAP DETECTION)
For the study_characteristics, you MUST extract these fields for cohort tracking:

1. **enrollment_period_start** and **enrollment_period_end**: 
   - Look for "patients enrolled between...", "study period", "from X to Y"
   - Format as "YYYY-MM" or "YYYY" if only year available
   - Example: "2018-01" to "2020-12"

2. **centers**: List of transplant centers/hospitals by name
   - Look in Methods section, typically "patients were enrolled at..."
   - Include city/country if center name alone is ambiguous
   - Example: ["University Hospital Zurich", "King's College London"]

3. **registry_id**: Clinical trial registration (NCT number, ISRCTN, etc.)
   - Look at end of abstract, methods, or supplementary
   - Include the full ID, e.g., "NCT02584283"

4. **countries**: All countries where patients were enrolled

These are essential for detecting overlapping patient cohorts between studies!

## IMPORTANT
- Use null for missing values
- Include source_location for all extracted data points
- For RCT studies, complete rob_rct in submit_extraction
- For NRS studies, complete rob_nrs in submit_extraction

Start by identifying the study design and calling get_rob_skill()."""

        contents.append(system_prompt)
        
        # Initial config with only get_rob_skill tool
        config = types.GenerateContentConfig(
            temperature=0,
            thinking_config=types.ThinkingConfig(
                thinking_level="medium"
            ),
            tools=INITIAL_TOOLS,  # Only get_rob_skill initially
        )
        
        # Track current study type for dynamic tool switching
        current_study_type = None
        extracted_study_id = "unknown"
        
        # Initialize extraction state for recording
        extraction_state = ExtractionState(
            study_id=extracted_study_id,  # Will be updated when identified
            model_name=self.model_name,
            pdf_paths=[str(p) for p in pdf_paths]
        ) if save_state else None
        
        # Record initial user content (system prompt + PDFs)
        if extraction_state:
            extraction_state.add_turn("user", [{"text": system_prompt}])
        
        # Initial call
        if verbose:
            print("  [Tool Calling] Starting extraction...")
        
        response = self.client.models.generate_content(
            model=self.model_name,
            contents=contents,
            config=config
        )
        
        # Track conversation history for multi-turn
        history = list(contents)
        
        # Process tool calls in a loop
        max_iterations = 5
        iteration = 0
        final_result = None
        
        while iteration < max_iterations:
            iteration += 1
            
            # Check for thinking content
            if verbose and hasattr(response, 'candidates') and response.candidates:
                for part in response.candidates[0].content.parts:
                    if hasattr(part, 'thought') and part.thought:
                        print(f"\n  💭 Thinking:\n{part.text[:500]}...")
            
            # Check for function calls
            function_calls = []
            if hasattr(response, 'candidates') and response.candidates:
                for part in response.candidates[0].content.parts:
                    if hasattr(part, 'function_call') and part.function_call:
                        function_calls.append(part.function_call)
            
            if not function_calls:
                # No more tool calls, check if we have text response
                if response.text:
                    try:
                        final_result = json.loads(response.text)
                    except json.JSONDecodeError:
                        pass
                break
            
            # Process each function call
            tool_responses = []
            for fc in function_calls:
                if verbose:
                    print(f"  🔧 Tool call: {fc.name}({json.dumps(fc.args, ensure_ascii=False)[:100]}...)")
                
                if fc.name == "get_rob_skill":
                    # JIT skill injection + dynamic tool switching
                    study_type = fc.args.get("study_type", "NRS")
                    study_id = fc.args.get("study_id", "Unknown")
                    skill = get_skill(study_type)
                    current_study_type = study_type
                    
                    if verbose:
                        print(f"  📥 Injecting {study_type} RoB skill for {study_id}")
                        print(f"  🔄 Switching to submit_extraction_{study_type.lower()} tool")
                    
                    # Update config with appropriate submit tool for next turn
                    config = types.GenerateContentConfig(
                        temperature=0,
                        thinking_config=types.ThinkingConfig(
                            thinking_level="medium"
                        ),
                        tools=get_extraction_tools(study_type),  # Dynamic tool based on study type
                    )
                    
                    tool_responses.append(
                        types.Part.from_function_response(
                            name="get_rob_skill",
                            response={"instructions": skill}
                        )
                    )
                    
                    # Inject RoB-related files (disclosure, protocol, COI) if available
                    if rob_files:
                        if verbose:
                            print(f"  📄 Injecting {len(rob_files)} RoB-related files")
                        for rob_path in rob_files:
                            try:
                                with open(rob_path, 'rb') as f:
                                    rob_data = f.read()
                                tool_responses.append(
                                    types.Part.from_bytes(
                                        data=rob_data,
                                        mime_type='application/pdf'
                                    )
                                )
                                if verbose:
                                    print(f"    • {Path(rob_path).name}")
                            except Exception as e:
                                if verbose:
                                    print(f"    ⚠️ Failed to inject {Path(rob_path).name}: {e}")
                    
                    # Record tool response in state
                    if extraction_state:
                        extraction_state.add_tool_response("get_rob_skill", {"instructions": "[SKILL INJECTED]"})
                    
                elif fc.name == "submit_extraction":
                    # Final extraction - return the data
                    if verbose:
                        print("  ✅ Extraction submitted")
                    final_result = fc.args
                    
                    # Update study_id in state if available
                    if extraction_state and final_result:
                        study_chars = final_result.get("study_characteristics", {})
                        extracted_study_id = study_chars.get("study_id", "unknown")
                        extraction_state.study_id = extracted_study_id
                        extraction_state.study_type = current_study_type
                    
                    # Continue to allow model to finish if needed
                    tool_responses.append(
                        types.Part.from_function_response(
                            name="submit_extraction",
                            response={"status": "success", "message": "Extraction data received"}
                        )
                    )
            
            # Record model response in state (including thought signatures)
            if extraction_state and response.candidates:
                extraction_state.add_model_response(response.candidates[0].content)
            
            if final_result and fc.name == "submit_extraction":
                # We have the final result, exit loop
                break
            
            # Add assistant response and tool responses to history
            history.append(response.candidates[0].content)
            history.append(types.Content(parts=tool_responses, role="user"))
            
            # Continue conversation
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=history,
                config=config
            )
        
        if final_result is None:
            print("  ⚠️ Warning: No final result from tool calling, using empty dict")
            final_result = {}
        
        # Record token usage from last response
        if extraction_state and response.usage_metadata:
            um = response.usage_metadata
            extraction_state.token_usage["prompt_tokens"] += um.prompt_token_count or 0
            extraction_state.token_usage["output_tokens"] += um.candidates_token_count or 0
            extraction_state.token_usage["thoughts_tokens"] += getattr(um, 'thoughts_token_count', 0) or 0
            extraction_state.token_usage["total_tokens"] += um.total_token_count or 0
        
        # Finalize and save extraction state
        if extraction_state:
            extraction_state.final_result = final_result
            if save_state:
                # Save to extraction_states directory
                states_dir = Path(__file__).parent / "extraction_states"
                states_dir.mkdir(exist_ok=True)
                
                if state_path is None:
                    state_path = str(states_dir / f"extraction_state_{extraction_state.study_id}.json")
                
                saved_path = extraction_state.save(state_path)
                if verbose:
                    print(f"  💾 Extraction state saved to: {saved_path}")
                    print(f"  📊 Token usage: {extraction_state.token_usage['total_tokens']:,} total")
        
        return final_result, extraction_state


# =============================================================================
# Excel Writer (동일한 구현 - 기존 코드에서 가져옴)
# =============================================================================

class ExcelWriter:
    def __init__(self, template_path: str, output_path: str):
        """Load template or existing output workbook"""
        
        # If output already exists, append to it
        if os.path.exists(output_path):
            print(f"  Appending to existing file: {output_path}")
            self.wb = openpyxl.load_workbook(output_path)
        else:
            # Copy template to output
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found: {template_path}")
            print(f"  Creating new file from template")
            self.wb = openpyxl.load_workbook(template_path)
        
        self.output_path = output_path
        
        # Map sheet keys to actual worksheets
        self.sheets = {}
        for key, name in SHEETS.items():
            if name in self.wb.sheetnames:
                self.sheets[key] = self.wb[name]
            else:
                print(f"  Warning: Sheet '{name}' not found in template")
    
    def save(self):
        """Save workbook"""
        self.wb.save(self.output_path)
        print(f"  Saved to: {self.output_path}")
    
    def _get_sheet(self, key: str):
        return self.sheets.get(key)
    
    def _append_row(self, key: str, data: List[Any]):
        """Append a row to specified sheet"""
        sheet = self._get_sheet(key)
        if sheet:
            sheet.append(data)
        else:
            print(f"  Warning: Could not write to sheet '{key}'")
    
    def write_study_characteristics(self, data: Dict[str, Any]):
        """Write to Study_Characteristics sheet"""
        # Handle enrollment period - can be nested or flat
        enrollment_start = None
        enrollment_end = None
        if isinstance(data.get("enrollment_period"), dict):
            enrollment_start = data.get("enrollment_period", {}).get("start")
            enrollment_end = data.get("enrollment_period", {}).get("end")
        else:
            enrollment_start = data.get("enrollment_period_start")
            enrollment_end = data.get("enrollment_period_end")
        
        # Handle centers - can be array or string
        centers = data.get("centers", [])
        if isinstance(centers, list):
            centers_str = ", ".join(centers)
        else:
            centers_str = centers
        
        row_data = [
            data.get("study_id"),
            data.get("first_author"),
            data.get("year"),
            data.get("title"),
            data.get("journal"),
            data.get("doi"),
            data.get("study_design"),
            data.get("is_multicenter"),
            centers_str,  # Added Centers column
            ", ".join(data.get("countries", [])) if isinstance(data.get("countries"), list) else data.get("countries"),
            enrollment_start,
            enrollment_end,
            data.get("registry_id"),
            data.get("intervention_type"),
            data.get("comparator"),
            data.get("n_intervention"),
            data.get("n_control"),
            data.get("n_total"),
            data.get("donor_type"),
            data.get("ecd_definition_used"),
            data.get("ecd_percentage"),
            data.get("donor_age", {}).get("intervention") if isinstance(data.get("donor_age"), dict) else None,
            data.get("donor_age", {}).get("control") if isinstance(data.get("donor_age"), dict) else None,
            data.get("donor_bmi", {}).get("intervention") if isinstance(data.get("donor_bmi"), dict) else None,
            data.get("donor_bmi", {}).get("control") if isinstance(data.get("donor_bmi"), dict) else None,
            data.get("donor_risk_score", {}).get("score_type") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("donor_risk_score", {}).get("intervention") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("donor_risk_score", {}).get("control") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("recipient_age", {}).get("intervention") if isinstance(data.get("recipient_age"), dict) else None,
            data.get("recipient_age", {}).get("control") if isinstance(data.get("recipient_age"), dict) else None,
            data.get("meld_score", {}).get("intervention") if isinstance(data.get("meld_score"), dict) else None,
            data.get("meld_score", {}).get("control") if isinstance(data.get("meld_score"), dict) else None,
            data.get("matching_method"),
            ", ".join(data.get("matching_variables", [])) if isinstance(data.get("matching_variables"), list) else data.get("matching_variables"),
            data.get("notes")
        ]
        self._append_row("characteristics", row_data)
    
    def write_perfusion_settings(self, data: Dict[str, Any]):
        """Write to Perfusion_Settings sheet"""
        co_interventions = data.get("co_interventions", [])
        co_int_str = ""
        co_int_details = ""
        if co_interventions:
            if isinstance(co_interventions[0], dict):
                co_int_str = ", ".join([c.get("agent", "") for c in co_interventions])
                co_int_details = "; ".join([f"{c.get('agent')}: {c.get('dose', '')} ({c.get('timing', '')})" for c in co_interventions])
            else:
                co_int_str = ", ".join(co_interventions)
        
        viability = data.get("viability_assessment", {})
        
        row_data = [
            data.get("study_id"),
            data.get("device_name"),
            data.get("device_portable"),
            data.get("cannulation"),
            data.get("perfusate_type"),
            ", ".join(data.get("perfusate_additives", [])) if isinstance(data.get("perfusate_additives"), list) else data.get("perfusate_additives"),
            data.get("temperature_setting"),
            data.get("temperature_celsius"),
            data.get("oxygenation", {}).get("active") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("oxygenation", {}).get("pO2_target") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("oxygenation", {}).get("flow_rate") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("pressure_settings", {}).get("portal_vein_mmHg") if isinstance(data.get("pressure_settings"), dict) else None,
            data.get("pressure_settings", {}).get("hepatic_artery_mmHg") if isinstance(data.get("pressure_settings"), dict) else None,
            co_int_str if co_int_str else None,
            co_int_details if co_int_details else None,
            viability.get("performed") if isinstance(viability, dict) else None,
            ", ".join(viability.get("criteria_used", [])) if isinstance(viability, dict) and isinstance(viability.get("criteria_used"), list) else None,
            viability.get("discard_rate_intervention") if isinstance(viability, dict) else None,
            viability.get("discard_rate_control") if isinstance(viability, dict) else None,
            data.get("notes")
        ]
        self._append_row("perfusion", row_data)
    
    def write_time_metrics(self, data: Dict[str, Any]):
        """Write to Time_Metrics sheet"""
        row_data = [
            data.get("study_id"),
            data.get("perfusion_initiation"),
            data.get("functional_wit_minutes", {}).get("intervention") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("functional_wit_minutes", {}).get("control") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("functional_wit_minutes", {}).get("measure") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("intervention") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("control") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("measure") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("perfusion_time_hours", {}).get("value") if isinstance(data.get("perfusion_time_hours"), dict) else data.get("perfusion_time_hours"),
            data.get("perfusion_time_hours", {}).get("measure") if isinstance(data.get("perfusion_time_hours"), dict) else None,
            data.get("perfusion_time_hours", {}).get("range_or_sd") if isinstance(data.get("perfusion_time_hours"), dict) else None,
            data.get("total_preservation_time_hours", {}).get("intervention") if isinstance(data.get("total_preservation_time_hours"), dict) else None,
            data.get("total_preservation_time_hours", {}).get("control") if isinstance(data.get("total_preservation_time_hours"), dict) else None,
            data.get("perfusion_to_preservation_ratio"),
            data.get("long_short_classification"),
            data.get("notes")
        ]
        self._append_row("time", row_data)
    
    def write_outcome_data(self, data: Dict[str, Any]):
        """Write to Outcome_Data sheet (binary outcomes)"""
        study_id = data.get("study_id")
        follow_up = data.get("follow_up_months")
        
        # Build row - matching template column order
        row_data = [study_id, follow_up]
        
        # EAD (12 fields - includes source)
        ead = data.get("ead", {})
        row_data.extend([
            ead.get("reported"), ead.get("definition"),
            ead.get("intervention_events"), ead.get("intervention_total"),
            ead.get("control_events"), ead.get("control_total"),
            ead.get("rr") or ead.get("effect_estimate"),
            ead.get("ci_lower"), ead.get("ci_upper"), ead.get("p_value"),
            ead.get("source_quote"), ead.get("source_location")
        ])
        
        # NAS (13 fields - includes follow_up_for_nas and source)
        nas = data.get("nas", {})
        row_data.extend([
            nas.get("reported"), nas.get("definition"), nas.get("follow_up_for_nas_months"),
            nas.get("intervention_events"), nas.get("intervention_total"),
            nas.get("control_events"), nas.get("control_total"),
            nas.get("rr") or nas.get("effect_estimate"),
            nas.get("ci_lower"), nas.get("ci_upper"), nas.get("p_value"),
            nas.get("source_quote"), nas.get("source_location")
        ])
        
        # TBC (12 fields - includes source)
        tbc = data.get("tbc", {})
        row_data.extend([
            tbc.get("reported"), tbc.get("definition"),
            tbc.get("intervention_events"), tbc.get("intervention_total"),
            tbc.get("control_events"), tbc.get("control_total"),
            tbc.get("rr") or tbc.get("effect_estimate"),
            tbc.get("ci_lower"), tbc.get("ci_upper"), tbc.get("p_value"),
            tbc.get("source_quote"), tbc.get("source_location")
        ])
        
        # Major Complications (12 fields - includes source)
        mc = data.get("major_complications", {})
        row_data.extend([
            mc.get("reported"), mc.get("definition"),
            mc.get("intervention_events"), mc.get("intervention_total"),
            mc.get("control_events"), mc.get("control_total"),
            mc.get("rr") or mc.get("effect_estimate"),
            mc.get("ci_lower"), mc.get("ci_upper"), mc.get("p_value"),
            mc.get("source_quote"), mc.get("source_location")
        ])
        
        # ACR (12 fields - includes source)
        acr = data.get("acr", {})
        row_data.extend([
            acr.get("reported"), acr.get("definition"),
            acr.get("intervention_events"), acr.get("intervention_total"),
            acr.get("control_events"), acr.get("control_total"),
            acr.get("rr") or acr.get("effect_estimate"),
            acr.get("ci_lower"), acr.get("ci_upper"), acr.get("p_value"),
            acr.get("source_quote"), acr.get("source_location")
        ])
        
        # PNF (12 fields - includes source)
        pnf = data.get("pnf", {})
        row_data.extend([
            pnf.get("reported"), pnf.get("definition"),
            pnf.get("intervention_events"), pnf.get("intervention_total"),
            pnf.get("control_events"), pnf.get("control_total"),
            pnf.get("rr") or pnf.get("effect_estimate"),
            pnf.get("ci_lower"), pnf.get("ci_upper"), pnf.get("p_value"),
            pnf.get("source_quote"), pnf.get("source_location")
        ])
        
        # HAT (11 fields - no definition, includes source)
        hat = data.get("hat", {})
        row_data.extend([
            hat.get("reported"),
            hat.get("intervention_events"), hat.get("intervention_total"),
            hat.get("control_events"), hat.get("control_total"),
            hat.get("rr") or hat.get("effect_estimate"),
            hat.get("ci_lower"), hat.get("ci_upper"), hat.get("p_value"),
            hat.get("source_quote"), hat.get("source_location")
        ])
        
        # Retransplantation (12 fields - includes source)
        retx = data.get("retransplantation", {})
        row_data.extend([
            retx.get("reported"), retx.get("timeframe"),
            retx.get("intervention_events"), retx.get("intervention_total"),
            retx.get("control_events"), retx.get("control_total"),
            retx.get("rr") or retx.get("effect_estimate"),
            retx.get("ci_lower"), retx.get("ci_upper"), retx.get("p_value"),
            retx.get("source_quote"), retx.get("source_location")
        ])
        
        # AKI (12 fields - includes source)
        aki = data.get("aki", {})
        row_data.extend([
            aki.get("reported"), aki.get("definition"),
            aki.get("intervention_events"), aki.get("intervention_total"),
            aki.get("control_events"), aki.get("control_total"),
            aki.get("rr") or aki.get("effect_estimate"),
            aki.get("ci_lower"), aki.get("ci_upper"), aki.get("p_value"),
            aki.get("source_quote"), aki.get("source_location")
        ])
        
        # RRT (11 fields - no definition, includes source)
        rrt = data.get("rrt", {})
        row_data.extend([
            rrt.get("reported"),
            rrt.get("intervention_events"), rrt.get("intervention_total"),
            rrt.get("control_events"), rrt.get("control_total"),
            rrt.get("rr") or rrt.get("effect_estimate"),
            rrt.get("ci_lower"), rrt.get("ci_upper"), rrt.get("p_value"),
            rrt.get("source_quote"), rrt.get("source_location")
        ])
        
        # PRS (12 fields - includes source)
        prs = data.get("prs", {})
        row_data.extend([
            prs.get("reported"), prs.get("definition"),
            prs.get("intervention_events"), prs.get("intervention_total"),
            prs.get("control_events"), prs.get("control_total"),
            prs.get("rr") or prs.get("effect_estimate"),
            prs.get("ci_lower"), prs.get("ci_upper"), prs.get("p_value"),
            prs.get("source_quote"), prs.get("source_location")
        ])
        
        # Graft Survival by timepoint (6 fields each: Reported, Int_Events, Int_Total, Ctrl_Events, Ctrl_Total, Source)
        gs = data.get("graft_survival", {})
        for tp in ["30d", "90d", "1yr", "3yr", "5yr"]:
            tp_data = gs.get(tp, {})
            row_data.extend([
                tp_data.get("reported"),
                tp_data.get("intervention_events"), tp_data.get("intervention_total"),
                tp_data.get("control_events"), tp_data.get("control_total"),
                tp_data.get("source_location")
            ])
        
        # Patient Survival by timepoint (6 fields each)
        ps = data.get("patient_survival", {})
        for tp in ["30d", "90d", "1yr", "3yr", "5yr"]:
            tp_data = ps.get(tp, {})
            row_data.extend([
                tp_data.get("reported"),
                tp_data.get("intervention_events"), tp_data.get("intervention_total"),
                tp_data.get("control_events"), tp_data.get("control_total"),
                tp_data.get("source_location")
            ])
        
        self._append_row("outcomes", row_data)
    
    def write_continuous_outcomes(self, data: Dict[str, Any]):
        """Write to Continuous_Outcomes sheet"""
        study_id = data.get("study_id")
        
        # Hospital Stay
        hosp = data.get("hospital_stay_days", {})
        # ICU Stay
        icu = data.get("icu_stay_days", {})
        # Graft Survival
        graft = data.get("graft_survival_1yr", {})
        # Patient Survival
        pat = data.get("patient_survival_1yr", {})
        # Utilization
        util = data.get("utilization_rate", {})
        
        row_data = [
            study_id,
            # Hospital Stay (9 fields including source)
            hosp.get("reported") if isinstance(hosp, dict) else None,
            hosp.get("intervention_mean_or_median") if isinstance(hosp, dict) else None,
            hosp.get("intervention_sd_or_iqr") if isinstance(hosp, dict) else None,
            hosp.get("control_mean_or_median") if isinstance(hosp, dict) else None,
            hosp.get("control_sd_or_iqr") if isinstance(hosp, dict) else None,
            hosp.get("measure") if isinstance(hosp, dict) else None,
            hosp.get("p_value") if isinstance(hosp, dict) else None,
            hosp.get("source_quote") if isinstance(hosp, dict) else None,
            hosp.get("source_location") if isinstance(hosp, dict) else None,
            # ICU Stay (9 fields including source)
            icu.get("reported") if isinstance(icu, dict) else None,
            icu.get("intervention_mean_or_median") if isinstance(icu, dict) else None,
            icu.get("intervention_sd_or_iqr") if isinstance(icu, dict) else None,
            icu.get("control_mean_or_median") if isinstance(icu, dict) else None,
            icu.get("control_sd_or_iqr") if isinstance(icu, dict) else None,
            icu.get("measure") if isinstance(icu, dict) else None,
            icu.get("p_value") if isinstance(icu, dict) else None,
            icu.get("source_quote") if isinstance(icu, dict) else None,
            icu.get("source_location") if isinstance(icu, dict) else None,
            # Graft Survival 1yr (13 fields including source)
            graft.get("reported") if isinstance(graft, dict) else None,
            graft.get("intervention_percent") if isinstance(graft, dict) else None,
            graft.get("control_percent") if isinstance(graft, dict) else None,
            graft.get("intervention_events") if isinstance(graft, dict) else None,
            graft.get("intervention_total") if isinstance(graft, dict) else None,
            graft.get("control_events") if isinstance(graft, dict) else None,
            graft.get("control_total") if isinstance(graft, dict) else None,
            graft.get("hr") if isinstance(graft, dict) else None,
            graft.get("ci_lower") if isinstance(graft, dict) else None,
            graft.get("ci_upper") if isinstance(graft, dict) else None,
            graft.get("p_value") if isinstance(graft, dict) else None,
            graft.get("source_quote") if isinstance(graft, dict) else None,
            graft.get("source_location") if isinstance(graft, dict) else None,
            # Patient Survival 1yr (13 fields including source)
            pat.get("reported") if isinstance(pat, dict) else None,
            pat.get("intervention_percent") if isinstance(pat, dict) else None,
            pat.get("control_percent") if isinstance(pat, dict) else None,
            pat.get("intervention_events") if isinstance(pat, dict) else None,
            pat.get("intervention_total") if isinstance(pat, dict) else None,
            pat.get("control_events") if isinstance(pat, dict) else None,
            pat.get("control_total") if isinstance(pat, dict) else None,
            pat.get("hr") if isinstance(pat, dict) else None,
            pat.get("ci_lower") if isinstance(pat, dict) else None,
            pat.get("ci_upper") if isinstance(pat, dict) else None,
            pat.get("p_value") if isinstance(pat, dict) else None,
            pat.get("source_quote") if isinstance(pat, dict) else None,
            pat.get("source_location") if isinstance(pat, dict) else None,
            # Utilization (7 fields including source)
            util.get("reported") if isinstance(util, dict) else None,
            util.get("intervention_utilized") if isinstance(util, dict) else None,
            util.get("intervention_offered") if isinstance(util, dict) else None,
            util.get("control_utilized") if isinstance(util, dict) else None,
            util.get("control_offered") if isinstance(util, dict) else None,
            util.get("source_quote") if isinstance(util, dict) else None,
            util.get("source_location") if isinstance(util, dict) else None,
            data.get("notes")
        ]
        self._append_row("continuous", row_data)
    
    def write_rob2(self, data: Dict[str, Any]):
        """Write to RoB2_RCT sheet"""
        d1 = data.get("d1_randomization", {})
        d2 = data.get("d2_deviations", {})
        d3 = data.get("d3_missing_data", {})
        d4 = data.get("d4_measurement", {})
        d5 = data.get("d5_selection", {})
        
        def get_quotes(domain_data):
            quotes = domain_data.get("support_quotes", [])
            if isinstance(quotes, list):
                return "; ".join([q.get("quote", q) if isinstance(q, dict) else str(q) for q in quotes])
            return str(quotes) if quotes else None
        
        def get_sq(domain_data, key):
            sq = domain_data.get("signaling_questions", {})
            return sq.get(key) if isinstance(sq, dict) else None
        
        row_data = [
            data.get("study_id"),
            # D1
            d1.get("judgment"),
            get_sq(d1, "random_sequence_generation"),
            get_sq(d1, "allocation_concealment"),
            get_sq(d1, "baseline_imbalances"),
            get_quotes(d1),
            # D2
            d2.get("judgment"),
            get_sq(d2, "participants_blinded"),
            get_sq(d2, "personnel_blinded"),
            get_sq(d2, "intention_to_treat"),
            get_sq(d2, "deviations_occurred"),
            get_quotes(d2),
            # D3
            d3.get("judgment"),
            get_sq(d3, "outcome_data_available"),
            get_sq(d3, "dropout_rate_intervention"),
            get_sq(d3, "dropout_rate_control"),
            get_sq(d3, "dropout_reasons_balanced"),
            get_quotes(d3),
            # D4
            d4.get("judgment"),
            get_sq(d4, "outcome_assessor_blinded"),
            get_sq(d4, "outcome_objective"),
            get_quotes(d4),
            # D5
            d5.get("judgment"),
            get_sq(d5, "preregistered_protocol"),
            data.get("registry_id"),
            get_sq(d5, "outcomes_match_protocol"),
            get_quotes(d5),
            # Overall
            data.get("overall_judgment"),
            data.get("overall_rationale")
        ]
        self._append_row("rob2", row_data)
    
    def write_robins_i(self, data: Dict[str, Any]):
        """Write to ROBINS_I_NRS sheet"""
        d1 = data.get("d1_confounding", {})
        d2 = data.get("d2_selection", {})
        d3 = data.get("d3_classification", {})
        d4 = data.get("d4_deviations", {})
        d5 = data.get("d5_missing_data", {})
        d6 = data.get("d6_measurement", {})
        d7 = data.get("d7_selection", {})
        
        def get_quotes(domain_data):
            quotes = domain_data.get("support_quotes", [])
            if isinstance(quotes, list):
                return "; ".join([q.get("quote", q) if isinstance(q, dict) else str(q) for q in quotes])
            return str(quotes) if quotes else None
        
        def get_sq(domain_data, key):
            sq = domain_data.get("signaling_questions", {})
            return sq.get(key) if isinstance(sq, dict) else None
        
        # Handle confounders list
        confounders = get_sq(d1, "confounders_considered")
        if isinstance(confounders, list):
            confounders = ", ".join(confounders)
        
        row_data = [
            data.get("study_id"),
            # D1
            d1.get("judgment"),
            confounders,
            get_sq(d1, "adjustment_method"),
            get_sq(d1, "residual_confounding_likely"),
            get_quotes(d1),
            # D2
            d2.get("judgment"),
            get_sq(d2, "selection_into_study"),
            get_sq(d2, "exclusions_post_intervention"),
            get_quotes(d2),
            # D3
            d3.get("judgment"),
            get_sq(d3, "classification_based_on"),
            get_sq(d3, "misclassification_possible"),
            get_quotes(d3),
            # D4
            d4.get("judgment"),
            get_sq(d4, "co_interventions_balanced"),
            get_sq(d4, "switches_between_groups"),
            get_quotes(d4),
            # D5
            d5.get("judgment"),
            get_sq(d5, "data_available_for_all"),
            get_sq(d5, "differential_missingness"),
            get_quotes(d5),
            # D6
            d6.get("judgment"),
            get_sq(d6, "outcome_assessors_aware"),
            get_sq(d6, "outcome_ascertainment_comparable"),
            get_quotes(d6),
            # D7
            d7.get("judgment"),
            get_sq(d7, "prespecified_analysis"),
            get_sq(d7, "multiple_analyses_performed"),
            get_quotes(d7),
            # Overall
            data.get("overall_judgment"),
            data.get("overall_rationale")
        ]
        self._append_row("robins", row_data)
    
    def write_notes(self, data: Dict[str, Any], study_id: str, extractor: str = "Gemini"):
        """Write to Extraction_Notes sheet"""
        concerns = data.get("data_quality_concerns", [])
        items = data.get("unclear_items_for_review", [])
        overlaps = data.get("potential_overlaps", [])
        
        overlap_study = ""
        overlap_reason = ""
        if overlaps and isinstance(overlaps, list) and len(overlaps) > 0:
            if isinstance(overlaps[0], dict):
                overlap_study = overlaps[0].get("study_id", "")
                overlap_reason = overlaps[0].get("overlap_reason", "")
            else:
                overlap_study = str(overlaps[0])
        
        row_data = [
            study_id,
            datetime.now().strftime("%Y-%m-%d"),
            extractor,
            "; ".join(concerns) if isinstance(concerns, list) else concerns,
            "; ".join(items) if isinstance(items, list) else items,
            overlap_study,
            overlap_reason,
            data.get("eligibility_status"),
            data.get("exclusion_reason"),
            data.get("general_notes")
        ]
        self._append_row("notes", row_data)


# =============================================================================
# Cohort Registry Writer (cohort_tracking.xlsx 자동 업데이트)
# =============================================================================

class CohortRegistryWriter:
    """Auto-populate Study Registry in cohort_tracking.xlsx from extraction results"""
    
    def __init__(self, cohort_tracking_path: str = None):
        if cohort_tracking_path is None:
            # Use absolute path relative to this script
            script_dir = Path(__file__).parent
            cohort_tracking_path = str(script_dir.parent / "Papers" / "cohort_tracking.xlsx")
        
        self.cohort_path = cohort_tracking_path
        if os.path.exists(cohort_tracking_path):
            try:
                self.wb = openpyxl.load_workbook(cohort_tracking_path)
            except Exception as e:
                print(f"  Warning: Could not load cohort_tracking.xlsx: {e}")
                self.wb = None
        else:
            print(f"  Warning: cohort_tracking.xlsx not found at {cohort_tracking_path}")
            self.wb = None
    
    def save(self):
        if self.wb:
            self.wb.save(self.cohort_path)
            print(f"  Saved cohort tracking to: {self.cohort_path}")
    
    def add_to_registry(self, data: Dict[str, Any]):
        """Add study to Study Registry sheet"""
        if not self.wb or "Study Registry" not in self.wb.sheetnames:
            return
        
        ws = self.wb["Study Registry"]
        
        # Check if already exists
        study_id = data.get("study_id")
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value == study_id:
                print(f"  Study {study_id} already in registry, skipping")
                return
        
        # Format enrollment period
        enrollment = ""
        if data.get("enrollment_period_start") and data.get("enrollment_period_end"):
            enrollment = f"{data.get('enrollment_period_start')}/{data.get('enrollment_period_end')}"
        elif data.get("enrollment_period_start"):
            enrollment = data.get("enrollment_period_start")
        
        # Format centers
        centers = data.get("centers", [])
        if isinstance(centers, list):
            centers = ", ".join(centers)
        
        # Format countries
        countries = data.get("countries", [])
        if isinstance(countries, list):
            countries = ",".join(countries)
        
        # Build row - matching Study Registry columns
        row_data = [
            study_id,                          # Study_ID
            data.get("first_author"),          # First Author
            data.get("year"),                  # Year
            data.get("title", "")[:50],        # Title (truncated)
            data.get("journal"),               # Journal
            data.get("doi") or data.get("registry_id"),  # DOI/PMID
            data.get("study_design"),          # Study Design
            centers,                           # Center(s)
            countries,                         # Country
            enrollment,                        # Enrollment Period
            data.get("registry_id"),           # Registry ID (NCT)
            None,                              # Parent Study ID
            None,                              # Relationship Type
            data.get("n_total"),               # Sample Size (Total)
            data.get("n_intervention"),        # Sample Size (Intervention)
            data.get("n_control"),             # Sample Size (Control)
            data.get("intervention_type"),     # Intervention Type
            None                               # Notes
        ]
        
        ws.append(row_data)
        print(f"  Added {study_id} to Study Registry")
    
    def check_overlaps(self, data: Dict[str, Any]) -> List[Dict]:
        """Check for potential cohort overlaps based on centers and enrollment period"""
        if not self.wb or "Study Registry" not in self.wb.sheetnames:
            return []
        
        ws = self.wb["Study Registry"]
        current_id = data.get("study_id")
        current_centers = set(data.get("centers", []) if isinstance(data.get("centers"), list) else [])
        current_nct = data.get("registry_id")
        
        overlaps = []
        
        for row in ws.iter_rows(min_row=2, max_col=18, values_only=True):
            if not row[0] or row[0] == current_id:
                continue
            
            other_id = row[0]
            other_centers = set(str(row[7] or "").split(", "))
            other_nct = row[10]
            
            # Check NCT match
            if current_nct and other_nct and current_nct == other_nct:
                overlaps.append({
                    "study_id": other_id,
                    "type": "Same NCT",
                    "evidence": f"Both use {current_nct}"
                })
            # Check center overlap
            elif current_centers & other_centers:
                shared = current_centers & other_centers
                overlaps.append({
                    "study_id": other_id,
                    "type": "Shared Centers",
                    "evidence": f"Both include: {', '.join(shared)}"
                })
        
        return overlaps


# =============================================================================
# Main Extraction Pipeline (최적화 버전: Quick Screen 제거)
# =============================================================================

class DataExtractorGemini:
    def __init__(self, model: str = DEFAULT_MODEL, template_path: str = TEMPLATE_PATH):
        self.llm = GeminiClient(model)
        self.template_path = template_path
        print(f"  Using model: {model}")
    
    def read_paper(self, file_path: str) -> str:
        """Read paper content from file (txt or PDF)"""
        path = Path(file_path)
        
        if path.suffix.lower() == '.pdf':
            print(f"  Reading PDF: {file_path}")
            text = extract_text_from_pdf(file_path)
            if len(text.strip()) < 500:
                print("  Trying OCR fallback...")
                text = extract_text_with_ocr_fallback(file_path)
            return text
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
    
    def full_extraction(self, paper_content: str) -> Dict[str, Any]:
        """Run full data extraction with structured schema (text mode)"""
        print("  Running full extraction (text mode)...")
        prompt = get_data_extraction_prompt(paper_content)
        return self.llm.extract(prompt, schema=FULL_EXTRACTION_SCHEMA)
    
    def full_extraction_pdf(self, pdf_paths: List[str]) -> Dict[str, Any]:
        """Run full data extraction with PDF direct input (multimodal)"""
        print("  Running full extraction (PDF direct mode)...")
        # Use a simpler prompt since PDF is provided directly
        prompt = """You are a data extraction expert for systematic reviews of Machine Perfusion in liver transplantation.

Analyze the provided PDF document(s) and extract ALL relevant data according to the schema.

Pay special attention to:
1. Tables with outcome data (events/totals for each group)
2. Figures with survival curves
3. Supplementary appendix tables
4. Exact definitions used for outcomes (EAD, NAS, etc.)

Extract data comprehensively and cite the source location for each data point."""
        
        return self.llm.extract_from_pdf(pdf_paths, prompt, schema=FULL_EXTRACTION_SCHEMA)
    
    def extract_rob(self, paper_content: str, study_type: str) -> Dict[str, Any]:
        """Run RoB-focused extraction with structured schema (text mode)"""
        print(f"  Extracting RoB information ({study_type}, text mode)...")
        prompt = get_rob_extraction_prompt(paper_content, study_type)
        schema = ROB2_EXTRACTION_SCHEMA if study_type == "RCT" else ROBINS_I_EXTRACTION_SCHEMA
        return self.llm.extract(prompt, schema=schema)
    
    def extract_rob_pdf(self, pdf_paths: List[str], study_type: str) -> Dict[str, Any]:
        """Run RoB-focused extraction with PDF direct input (multimodal)"""
        print(f"  Extracting RoB information ({study_type}, PDF direct mode)...")
        
        if study_type == "RCT":
            prompt = """Assess Risk of Bias using the RoB 2 tool for RCTs.

For each domain, provide:
- judgment: Low, Some concerns, or High
- rationale: Detailed explanation
- support_quotes: Direct quotes from the paper

Domains to assess:
1. Randomization Process
2. Deviations from Intended Interventions
3. Missing Outcome Data
4. Measurement of Outcome
5. Selection of Reported Result

Also provide overall_judgment and overall_rationale."""
            schema = ROB2_EXTRACTION_SCHEMA
        else:
            prompt = """Assess Risk of Bias using the ROBINS-I tool for non-randomized studies.

For each domain, provide:
- judgment: Low, Moderate, Serious, or Critical
- rationale: Detailed explanation

Domains to assess:
1. Confounding
2. Selection of Participants
3. Classification of Interventions
4. Deviations from Intended Interventions
5. Missing Data
6. Measurement of Outcomes
7. Selection of Reported Result

Also provide overall_judgment and overall_rationale."""
            schema = ROBINS_I_EXTRACTION_SCHEMA
        
        return self.llm.extract_from_pdf(pdf_paths, prompt, schema=schema)
    
    def process_paper(self, file_path: str, output_path: str, 
                      include_rob: bool = True,
                      supplementary_files: List[str] = None,
                      rob_files: List[str] = None,
                      pdf_direct: bool = False,
                      tool_calling: bool = False,
                      verbose: bool = False):
        """Process a single paper and write to Excel
        
        Args:
            file_path: Main paper PDF/text file
            output_path: Output Excel file
            include_rob: Include RoB extraction (ignored if tool_calling=True)
            supplementary_files: List of supplementary PDF/text files to include
            rob_files: List of RoB-related files (disclosure, protocol, COI) for injection during RoB assessment
            pdf_direct: If True, send PDFs directly to Gemini (multimodal)
            tool_calling: If True, use single API call with tool calling (recommended)
            verbose: If True, print detailed extraction progress
        """
        
        # Tool Calling Mode - Single API call with JIT skill injection
        if tool_calling:
            if not file_path.lower().endswith('.pdf'):
                raise ValueError("Tool calling mode requires PDF input")
            
            pdf_files = [file_path]
            if supplementary_files:
                pdf_files.extend([f for f in supplementary_files if f.lower().endswith('.pdf')])
            
            print(f"  Tool Calling Mode: {len(pdf_files)} file(s)")
            for pf in pdf_files:
                print(f"    • {Path(pf).name}")
            
            # Single API call with tools
            print("  [1/1] Unified extraction with tool calling...")
            result, extraction_state = self.llm.extract_with_tools(
                pdf_files, 
                rob_files=rob_files,
                verbose=verbose, 
                save_state=True
            )
            
            # Extract study info
            if "study_characteristics" in result and isinstance(result["study_characteristics"], dict):
                study_id = result["study_characteristics"].get("study_id", "Unknown")
                design = result["study_characteristics"].get("study_design", "")
            else:
                study_id = result.get("study_id", "Unknown")
                design = result.get("study_design", "")
            
            study_type = "NRS" if design and "RCT" not in design.upper() and "RANDOM" not in design.upper() else "RCT"
            print(f"        → Study: {study_id}")
            print(f"        → Design: {design} → {study_type}")
            
            # Prepare for writing - unified result contains everything
            full_result = result
            rob_result = result.get("rob_rct") if study_type == "RCT" else result.get("rob_nrs")
            
            # Write to Excel
            print(f"  Writing to Excel...")
            writer = ExcelWriter(self.template_path, output_path)
            
            if "study_characteristics" in result and isinstance(result["study_characteristics"], dict):
                writer.write_study_characteristics(result["study_characteristics"])
            
            if "perfusion_settings" in result and isinstance(result.get("perfusion_settings"), dict):
                perf = result["perfusion_settings"]
                perf["study_id"] = study_id
                writer.write_perfusion_settings(perf)
            
            if "time_metrics" in result and isinstance(result.get("time_metrics"), dict):
                time_data = result["time_metrics"]
                time_data["study_id"] = study_id
                writer.write_time_metrics(time_data)
            
            if "outcome_data" in result and isinstance(result.get("outcome_data"), dict):
                outcomes = result["outcome_data"]
                outcomes["study_id"] = study_id
                writer.write_outcome_data(outcomes)
                writer.write_continuous_outcomes(outcomes)
            
            if rob_result and isinstance(rob_result, dict):
                rob_result["study_id"] = study_id
                if study_type == "RCT":
                    writer.write_rob2(rob_result)
                else:
                    writer.write_robins_i(rob_result)
            
            notes = result.get("extraction_notes", {}) if isinstance(result, dict) else {}
            writer.write_notes(notes, study_id)
            
            writer.save()
            
            # Update Cohort Tracking
            if "study_characteristics" in result and isinstance(result["study_characteristics"], dict):
                try:
                    cohort_writer = CohortRegistryWriter()
                    cohort_writer.add_to_registry(result["study_characteristics"])
                    
                    # Check for overlaps
                    overlaps = cohort_writer.check_overlaps(result["study_characteristics"])
                    if overlaps:
                        print(f"  ⚠️ Potential overlaps detected:")
                        for ov in overlaps:
                            print(f"     - {ov['study_id']}: {ov['type']} ({ov['evidence']})")
                    
                    cohort_writer.save()
                except Exception as e:
                    print(f"  Warning: Could not update cohort tracking: {e}")
            
            return {
                "full_extraction": result,
                "rob_information": rob_result,
                "study_type": study_type,
                "study_id": study_id
            }
        
        # Legacy Mode - 2 API calls (PDF Direct or Text Mode)
        if pdf_direct and file_path.lower().endswith('.pdf'):
            # PDF Direct Mode - send files directly to Gemini
            pdf_files = [file_path]
            if supplementary_files:
                pdf_files.extend([f for f in supplementary_files if f.lower().endswith('.pdf')])
            
            print(f"  PDF Direct Mode: {len(pdf_files)} file(s)")
            for pf in pdf_files:
                print(f"    • {Path(pf).name}")
            
            # [1/2] Full extraction with PDF direct
            print("  [1/2] Full data extraction...")
            full_result = self.full_extraction_pdf(pdf_files)
            
            # Determine study type
            study_type = "RCT"
            if "study_characteristics" in full_result and isinstance(full_result["study_characteristics"], dict):
                design = full_result["study_characteristics"].get("study_design", "")
                study_id = full_result["study_characteristics"].get("study_id", "Unknown")
            else:
                design = full_result.get("study_design", full_result.get("design", ""))
                study_id = full_result.get("study_id", "Unknown")
            
            if design and "RCT" not in design.upper() and "RANDOM" not in design.upper():
                study_type = "NRS"
            
            print(f"        → Study: {study_id}")
            print(f"        → Design: {design} → {study_type}")
            
            # [2/2] RoB extraction with PDF direct
            rob_result = {}
            if include_rob:
                print(f"  [2/2] RoB extraction ({study_type})...")
                rob_result = self.extract_rob_pdf(pdf_files, study_type)
            else:
                print("  [2/2] RoB extraction skipped")
            
            paper_content = None  # Not used in PDF direct mode
        else:
            # Text Mode - extract text from PDFs first
            paper_content = self.read_paper(file_path)
            print(f"  Read {len(paper_content):,} characters from {Path(file_path).name}")
            
            # Read and append supplementary files
            if supplementary_files:
                for supp_file in supplementary_files:
                    try:
                        supp_content = self.read_paper(supp_file)
                        paper_content += f"\n\n{'='*60}\n"
                        paper_content += f"SUPPLEMENTARY MATERIAL: {Path(supp_file).name}\n"
                        paper_content += f"{'='*60}\n\n"
                        paper_content += supp_content
                        print(f"  + Added {len(supp_content):,} characters from {Path(supp_file).name}")
                    except Exception as e:
                        print(f"  ⚠️ Warning: Could not read {supp_file}: {e}")
                
                print(f"  Total content: {len(paper_content):,} characters")
            
            # [1/2] Full extraction (Quick Screen 제거)
            print("  [1/2] Full data extraction...")
            full_result = self.full_extraction(paper_content)
            
            # Determine study type for RoB from full extraction result
            study_type = "RCT"
            if "study_characteristics" in full_result:
                design = full_result["study_characteristics"].get("study_design", "")
                study_id = full_result["study_characteristics"].get("study_id", "Unknown")
            else:
                design = full_result.get("study_design", full_result.get("design", ""))
                study_id = full_result.get("study_id", "Unknown")
            
            if design and "RCT" not in design.upper() and "RANDOM" not in design.upper():
                study_type = "NRS"
            
            print(f"        → Study: {study_id}")
            print(f"        → Design: {design} → {study_type}")
            
            # [2/2] RoB extraction if requested
            rob_result = {}
            if include_rob:
                print(f"  [2/2] RoB extraction ({study_type})...")
                rob_result = self.extract_rob(paper_content, study_type)
            else:
                print("  [2/2] RoB extraction skipped")
        
        # Write to Excel
        print(f"  Writing to Excel...")
        writer = ExcelWriter(self.template_path, output_path)
        
        # Handle case where full_result is not a proper dict
        if not isinstance(full_result, dict):
            print(f"  ⚠️ Warning: Full extraction result is not a dict. Type: {type(full_result)}")
            full_result = {"raw_response": str(full_result), "parse_error": "Not a dict"}
        
        # Get study_id
        if "study_characteristics" in full_result and isinstance(full_result["study_characteristics"], dict):
            study_id = full_result["study_characteristics"].get("study_id")
            writer.write_study_characteristics(full_result["study_characteristics"])
        elif isinstance(full_result, dict):
            study_id = full_result.get("study_id")
            writer.write_study_characteristics(full_result)
        else:
            study_id = "Unknown"
        
        # Perfusion settings
        if "perfusion_settings" in full_result and isinstance(full_result.get("perfusion_settings"), dict):
            perf = full_result["perfusion_settings"]
            perf["study_id"] = study_id
            writer.write_perfusion_settings(perf)
        
        # Time metrics
        if "time_metrics" in full_result and isinstance(full_result.get("time_metrics"), dict):
            time_data = full_result["time_metrics"]
            time_data["study_id"] = study_id
            writer.write_time_metrics(time_data)
        
        # Outcomes
        if "outcome_data" in full_result and isinstance(full_result.get("outcome_data"), dict):
            outcomes = full_result["outcome_data"]
            outcomes["study_id"] = study_id
            writer.write_outcome_data(outcomes)
            writer.write_continuous_outcomes(outcomes)
        elif "outcomes" in full_result and isinstance(full_result.get("outcomes"), dict):
            outcomes = full_result["outcomes"]
            outcomes["study_id"] = study_id
            writer.write_outcome_data(outcomes)
            writer.write_continuous_outcomes(outcomes)
        
        # RoB
        if rob_result and isinstance(rob_result, dict):
            rob_result["study_id"] = study_id
            if study_type == "RCT":
                writer.write_rob2(rob_result)
            else:
                writer.write_robins_i(rob_result)
        elif rob_result:
            print(f"  ⚠️ Warning: RoB result is not a dict, skipping. Type: {type(rob_result)}")
        
        # Notes
        notes = full_result.get("extraction_notes", {}) if isinstance(full_result, dict) else {}
        writer.write_notes(notes, study_id)
        
        writer.save()
        
        return {
            "full_extraction": full_result,
            "rob_information": rob_result,
            "study_type": study_type,
            "study_id": study_id
        }
    
    def batch_process(self, input_dir: str, output_path: str):
        """Process all papers in a directory"""
        input_path = Path(input_dir)
        files = list(input_path.glob("*.pdf")) + list(input_path.glob("*.txt"))
        
        print(f"\nFound {len(files)} files to process")
        print(f"Output file: {output_path}")
        print(f"{'='*60}")
        
        results = []
        success_count = 0
        error_count = 0
        
        for i, file in enumerate(files, 1):
            print(f"\n[{i}/{len(files)}] Processing: {file.name}")
            print(f"-" * 40)
            try:
                result = self.process_paper(str(file), output_path, include_rob=True)
                results.append({
                    "file": file.name,
                    "success": True,
                    "study_id": result.get("study_id"),
                    "design": result.get("study_type"),
                    "result": result
                })
                success_count += 1
                print(f"  ✓ Success")
            except Exception as e:
                print(f"  ✗ Error: {e}")
                results.append({
                    "file": file.name,
                    "success": False,
                    "error": str(e)
                })
                error_count += 1
        
        # Summary
        print(f"\n{'='*60}")
        print(f"BATCH PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"Total files:    {len(files)}")
        print(f"Successful:     {success_count}")
        print(f"Errors:         {error_count}")
        print(f"Output file:    {output_path}")
        
        if success_count > 0:
            print(f"\nExtracted studies:")
            for r in results:
                if r.get("success"):
                    print(f"  • {r.get('study_id', 'Unknown')} ({r.get('design', '?')}) - {r['file']}")
        
        if error_count > 0:
            print(f"\nFailed files:")
            for r in results:
                if not r.get("success"):
                    print(f"  • {r['file']}: {r.get('error', 'Unknown error')}")
        
        return results


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Data Extractor (Gemini) for Machine Perfusion Systematic Review",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single paper extraction (RECOMMENDED: Tool Calling mode)
  python extractor_gemini.py -i paper.pdf -o results.xlsx --tool-calling
  
  # With verbose output to see thinking process
  python extractor_gemini.py -i paper.pdf -o results.xlsx --tool-calling --verbose
  
  # Legacy: 2 API calls mode (PDF direct)
  python extractor_gemini.py -i paper.pdf -o results.xlsx --pdf-direct
  
  # With supplementary materials
  python extractor_gemini.py -i paper.pdf -s supplementary.pdf -o results.xlsx --tool-calling
  
  # Batch process folder
  python extractor_gemini.py -i ./papers/ -o results.xlsx --batch
  
  # Save JSON alongside Excel
  python extractor_gemini.py -i paper.pdf -o results.xlsx --tool-calling --json

Environment Variables:
  GOOGLE_API_KEY or GEMINI_API_KEY: Your Gemini API key
        """
    )
    parser.add_argument("--input", "-i", required=True, help="Input file or directory")
    parser.add_argument("--supplementary", "-s", nargs="+", help="Supplementary file(s) to include")
    parser.add_argument("--output", "-o", default="extraction_results.xlsx", help="Output Excel file")
    parser.add_argument("--template", "-t", default=TEMPLATE_PATH, help="Template Excel file")
    parser.add_argument("--model", "-m", default=DEFAULT_MODEL, help="Gemini model to use")
    parser.add_argument("--no-rob", action="store_true", help="Skip RoB information extraction")
    parser.add_argument("--batch", action="store_true", help="Batch process all files in directory")
    parser.add_argument("--json", action="store_true", help="Also save results as JSON file")
    parser.add_argument("--pdf-direct", action="store_true", help="Send PDF directly to Gemini (multimodal, 2 API calls)")
    parser.add_argument("--tool-calling", action="store_true", help="Use tool calling mode (single API call, recommended)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed extraction progress and thinking")
    parser.add_argument("--update-tracking", metavar="FILE", help="Update cohort_tracking.xlsx with extraction results")
    
    args = parser.parse_args()
    
    print(f"\n{'='*60}")
    print(f"Data Extractor (Gemini) - Machine Perfusion Systematic Review")
    print(f"{'='*60}")
    print(f"Model: {args.model}")
    print(f"Template: {args.template}")
    print(f"Output: {args.output}")
    if args.tool_calling:
        print(f"Mode: Tool Calling (Single API call)")
    elif args.pdf_direct:
        print(f"Mode: PDF Direct (2 API calls)")
    else:
        print(f"Mode: Text (2 API calls)")
    if args.update_tracking:
        print(f"Tracking: {args.update_tracking}")
    
    extractor = DataExtractorGemini(model=args.model, template_path=args.template)
    
    if args.batch:
        results = extractor.batch_process(args.input, args.output)
        
        if args.json:
            json_path = args.output.replace('.xlsx', '_batch.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False, default=str)
            print(f"JSON saved to: {json_path}")
    else:
        # Handle supplementary files
        supplementary_files = args.supplementary if args.supplementary else []
        if supplementary_files:
            print(f"Supplementary files: {len(supplementary_files)}")
            for sf in supplementary_files:
                print(f"  • {sf}")
        
        result = extractor.process_paper(
            args.input, 
            args.output,
            include_rob=not args.no_rob,
            supplementary_files=supplementary_files,
            pdf_direct=args.pdf_direct,
            tool_calling=args.tool_calling,
            verbose=args.verbose
        )
        
        if args.json:
            json_path = args.output.replace('.xlsx', '.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False, default=str)
            print(f"JSON saved to: {json_path}")
        
        # Update cohort tracking if specified
        tracking_result = None
        if args.update_tracking:
            try:
                from cohort_tracker import CohortTracker
                tracker = CohortTracker(args.update_tracking)
                tracking_result = tracker.process_extraction(result)
                print(f"\n  📋 Cohort Tracking Updated:")
                print(f"     Registry: {tracking_result['registry']['status']}")
                print(f"     Outcome Matrix: {tracking_result['outcome_matrix']['status']}")
                if tracking_result.get('duplicates'):
                    print(f"     ⚠️ Potential duplicates found:")
                    for dup in tracking_result['duplicates']:
                        print(f"        - {dup['existing_study_id']} ({dup['match_type']})")
            except FileNotFoundError as e:
                print(f"  ⚠️ Tracking file not found: {args.update_tracking}")
            except Exception as e:
                print(f"  ⚠️ Error updating tracking: {e}")
        
        print(f"\n{'='*60}")
        print(f"EXTRACTION COMPLETE")
        print(f"{'='*60}")
        print(f"Study ID: {result.get('study_id')}")
        print(f"Study Type: {result.get('study_type')}")
        print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
