#!/usr/bin/env python3
"""
Data Extractor v2 for Machine Perfusion Systematic Review
목적: 메타분석용 데이터 수집 (eligibility 판단 X)
"""

import json
import os
import re
import shutil
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Local imports
from prompts_v2_data_collection import (
    get_data_extraction_prompt,
    get_quick_screen_prompt,
    get_rob_extraction_prompt,
    get_outcome_extraction_prompt
)
from pdf_utils import extract_text_from_pdf, extract_text_with_ocr_fallback


# =============================================================================
# Configuration
# =============================================================================

DEFAULT_MODEL = "claude-sonnet-4-5-20250929"
TEMPLATE_PATH = "data_extraction_template_v2.xlsx"

# Sheet names in template
SHEETS = {
    "characteristics": "Study_Characteristics",
    "perfusion": "Perfusion_Settings",
    "time": "Time_Metrics",
    "outcomes": "Outcome_Data",
    "continuous": "Continuous_Outcomes",
    "rob2": "RoB2_RCT",
    "robins": "ROBINS_I_NRS",
    "notes": "Extraction_Notes"
}


# =============================================================================
# LLM Client
# =============================================================================

class LLMClient:
    def __init__(self, model: str = DEFAULT_MODEL):
        self.client = anthropic.Anthropic()
        self.model = model
    
    def extract(self, prompt: str, max_tokens: int = 8000) -> Dict[str, Any]:
        """Call Claude API and parse JSON response"""
        message = self.client.messages.create(
            model=self.model,
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}]
        )
        
        text = message.content[0].text
        
        # Extract JSON from response
        json_match = re.search(r'\{[\s\S]*\}', text)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError as e:
                print(f"JSON parse error: {e}")
                return {"raw_response": text, "parse_error": str(e)}
        
        return {"raw_response": text}


# =============================================================================
# Excel Writer
# =============================================================================

class ExcelWriter:
    def __init__(self, template_path: str, output_path: str):
        """Load template or existing output workbook"""
        
        # If output already exists, append to it
        if os.path.exists(output_path):
            print(f"  Appending to existing file: {output_path}")
            self.wb = openpyxl.load_workbook(output_path)
        # Otherwise, copy from template
        elif os.path.exists(template_path):
            print(f"  Creating new file from template: {template_path}")
            shutil.copy(template_path, output_path)
            self.wb = openpyxl.load_workbook(output_path)
        else:
            # Try looking in script directory
            script_dir = Path(__file__).parent
            alt_template = script_dir / template_path
            if alt_template.exists():
                print(f"  Creating new file from template: {alt_template}")
                shutil.copy(alt_template, output_path)
                self.wb = openpyxl.load_workbook(output_path)
            else:
                raise FileNotFoundError(
                    f"Template not found at '{template_path}' or '{alt_template}'\n"
                    f"Run the template generator first or specify --template path"
                )
        
        self.output_path = output_path
        
        # Find next empty row for each sheet (skip header row)
        self.next_rows = {}
        for key, sheet_name in SHEETS.items():
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                # Find first empty row (max_row returns last used row)
                self.next_rows[key] = ws.max_row + 1 if ws.max_row > 1 else 2
    
    def save(self):
        """Save workbook"""
        self.wb.save(self.output_path)
        print(f"Saved to: {self.output_path}")
    
    def _get_sheet(self, key: str):
        return self.wb[SHEETS[key]]
    
    def _append_row(self, key: str, data: List[Any]):
        """Append a row to specified sheet"""
        ws = self._get_sheet(key)
        row = self.next_rows[key]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        self.next_rows[key] += 1
    
    def write_study_characteristics(self, data: Dict[str, Any]):
        """Write to Study_Characteristics sheet"""
        row_data = [
            data.get("study_id"),
            data.get("first_author"),
            data.get("year"),
            data.get("title"),
            data.get("journal"),
            data.get("doi"),
            data.get("study_design"),
            data.get("is_multicenter"),
            ", ".join(data.get("countries", [])) if isinstance(data.get("countries"), list) else data.get("countries"),
            data.get("enrollment_period", {}).get("start") if isinstance(data.get("enrollment_period"), dict) else None,
            data.get("enrollment_period", {}).get("end") if isinstance(data.get("enrollment_period"), dict) else None,
            data.get("registry_id"),
            data.get("intervention_type"),
            data.get("comparator"),
            data.get("n_intervention"),
            data.get("n_control"),
            data.get("n_total"),
            data.get("donor_type"),
            data.get("ecd_definition_used"),
            data.get("ecd_percentage"),
            data.get("donor_age", {}).get("intervention") if isinstance(data.get("donor_age"), dict) else None,
            data.get("donor_age", {}).get("control") if isinstance(data.get("donor_age"), dict) else None,
            data.get("donor_bmi", {}).get("intervention") if isinstance(data.get("donor_bmi"), dict) else None,
            data.get("donor_bmi", {}).get("control") if isinstance(data.get("donor_bmi"), dict) else None,
            data.get("donor_risk_score", {}).get("score_type") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("donor_risk_score", {}).get("intervention") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("donor_risk_score", {}).get("control") if isinstance(data.get("donor_risk_score"), dict) else None,
            data.get("recipient_age", {}).get("intervention") if isinstance(data.get("recipient_age"), dict) else None,
            data.get("recipient_age", {}).get("control") if isinstance(data.get("recipient_age"), dict) else None,
            data.get("meld_score", {}).get("intervention") if isinstance(data.get("meld_score"), dict) else None,
            data.get("meld_score", {}).get("control") if isinstance(data.get("meld_score"), dict) else None,
            data.get("matching_method"),
            ", ".join(data.get("matching_variables", [])) if isinstance(data.get("matching_variables"), list) else data.get("matching_variables"),
            data.get("notes")
        ]
        self._append_row("characteristics", row_data)
    
    def write_perfusion_settings(self, data: Dict[str, Any]):
        """Write to Perfusion_Settings sheet"""
        co_interventions = data.get("co_interventions", [])
        co_int_str = ""
        co_int_details = ""
        if co_interventions:
            if isinstance(co_interventions[0], dict):
                co_int_str = ", ".join([c.get("agent", "") for c in co_interventions])
                co_int_details = "; ".join([f"{c.get('agent')}: {c.get('dose', '')} ({c.get('timing', '')})" for c in co_interventions])
            else:
                co_int_str = ", ".join(co_interventions)
        
        viability = data.get("viability_assessment", {})
        
        row_data = [
            data.get("study_id"),
            data.get("device_name"),
            data.get("device_portable"),
            data.get("cannulation"),
            data.get("perfusate_type"),
            ", ".join(data.get("perfusate_additives", [])) if isinstance(data.get("perfusate_additives"), list) else data.get("perfusate_additives"),
            data.get("temperature_setting"),
            data.get("temperature_celsius"),
            data.get("oxygenation", {}).get("active") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("oxygenation", {}).get("pO2_target") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("oxygenation", {}).get("flow_rate") if isinstance(data.get("oxygenation"), dict) else None,
            data.get("pressure_settings", {}).get("portal_vein_mmHg") if isinstance(data.get("pressure_settings"), dict) else None,
            data.get("pressure_settings", {}).get("hepatic_artery_mmHg") if isinstance(data.get("pressure_settings"), dict) else None,
            co_int_str if co_int_str else None,
            co_int_details if co_int_details else None,
            viability.get("performed") if isinstance(viability, dict) else None,
            ", ".join(viability.get("criteria_used", [])) if isinstance(viability, dict) and isinstance(viability.get("criteria_used"), list) else None,
            viability.get("discard_rate_intervention") if isinstance(viability, dict) else None,
            viability.get("discard_rate_control") if isinstance(viability, dict) else None,
            data.get("notes")
        ]
        self._append_row("perfusion", row_data)
    
    def write_time_metrics(self, data: Dict[str, Any]):
        """Write to Time_Metrics sheet"""
        row_data = [
            data.get("study_id"),
            data.get("perfusion_initiation"),
            data.get("functional_wit_minutes", {}).get("intervention") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("functional_wit_minutes", {}).get("control") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("functional_wit_minutes", {}).get("measure") if isinstance(data.get("functional_wit_minutes"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("intervention") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("control") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("cold_ischemia_time_hours", {}).get("measure") if isinstance(data.get("cold_ischemia_time_hours"), dict) else None,
            data.get("perfusion_time_hours", {}).get("value") if isinstance(data.get("perfusion_time_hours"), dict) else data.get("perfusion_time_hours"),
            data.get("perfusion_time_hours", {}).get("measure") if isinstance(data.get("perfusion_time_hours"), dict) else None,
            data.get("perfusion_time_hours", {}).get("range_or_sd") if isinstance(data.get("perfusion_time_hours"), dict) else None,
            data.get("total_preservation_time_hours", {}).get("intervention") if isinstance(data.get("total_preservation_time_hours"), dict) else None,
            data.get("total_preservation_time_hours", {}).get("control") if isinstance(data.get("total_preservation_time_hours"), dict) else None,
            data.get("perfusion_to_preservation_ratio"),
            data.get("long_short_classification"),
            data.get("notes")
        ]
        self._append_row("time", row_data)
    
    def write_outcome_data(self, data: Dict[str, Any]):
        """Write to Outcome_Data sheet (binary outcomes)"""
        study_id = data.get("study_id")
        follow_up = data.get("follow_up_months")
        
        def get_outcome_fields(outcome_data: Dict) -> List:
            """Extract fields for one outcome"""
            if not outcome_data or not isinstance(outcome_data, dict):
                return [None] * 10  # 최대 필드 수
            return [
                outcome_data.get("reported"),
                outcome_data.get("definition"),
                outcome_data.get("intervention_events"),
                outcome_data.get("intervention_total"),
                outcome_data.get("control_events"),
                outcome_data.get("control_total"),
                outcome_data.get("rr") or outcome_data.get("effect_estimate"),
                outcome_data.get("ci_lower"),
                outcome_data.get("ci_upper"),
                outcome_data.get("p_value")
            ]
        
        # Build row - matching template column order
        row_data = [study_id, follow_up]
        
        # EAD (12 fields - includes source)
        ead = data.get("ead", {})
        row_data.extend([
            ead.get("reported"), ead.get("definition"),
            ead.get("intervention_events"), ead.get("intervention_total"),
            ead.get("control_events"), ead.get("control_total"),
            ead.get("rr") or ead.get("effect_estimate"),
            ead.get("ci_lower"), ead.get("ci_upper"), ead.get("p_value"),
            ead.get("source_quote"), ead.get("source_location")
        ])
        
        # NAS (13 fields - includes follow_up_for_nas and source)
        nas = data.get("nas", {})
        row_data.extend([
            nas.get("reported"), nas.get("definition"), nas.get("follow_up_for_nas_months"),
            nas.get("intervention_events"), nas.get("intervention_total"),
            nas.get("control_events"), nas.get("control_total"),
            nas.get("rr") or nas.get("effect_estimate"),
            nas.get("ci_lower"), nas.get("ci_upper"), nas.get("p_value"),
            nas.get("source_quote"), nas.get("source_location")
        ])
        
        # TBC (12 fields - includes source)
        tbc = data.get("tbc", {})
        row_data.extend([
            tbc.get("reported"), tbc.get("definition"),
            tbc.get("intervention_events"), tbc.get("intervention_total"),
            tbc.get("control_events"), tbc.get("control_total"),
            tbc.get("rr") or tbc.get("effect_estimate"),
            tbc.get("ci_lower"), tbc.get("ci_upper"), tbc.get("p_value"),
            tbc.get("source_quote"), tbc.get("source_location")
        ])
        
        # Major Complications (12 fields - includes source)
        mc = data.get("major_complications", {})
        row_data.extend([
            mc.get("reported"), mc.get("definition"),
            mc.get("intervention_events"), mc.get("intervention_total"),
            mc.get("control_events"), mc.get("control_total"),
            mc.get("rr") or mc.get("effect_estimate"),
            mc.get("ci_lower"), mc.get("ci_upper"), mc.get("p_value"),
            mc.get("source_quote"), mc.get("source_location")
        ])
        
        # ACR (12 fields - includes source)
        acr = data.get("acr", {})
        row_data.extend([
            acr.get("reported"), acr.get("definition"),
            acr.get("intervention_events"), acr.get("intervention_total"),
            acr.get("control_events"), acr.get("control_total"),
            acr.get("rr") or acr.get("effect_estimate"),
            acr.get("ci_lower"), acr.get("ci_upper"), acr.get("p_value"),
            acr.get("source_quote"), acr.get("source_location")
        ])
        
        # PNF (12 fields - includes source)
        pnf = data.get("pnf", {})
        row_data.extend([
            pnf.get("reported"), pnf.get("definition"),
            pnf.get("intervention_events"), pnf.get("intervention_total"),
            pnf.get("control_events"), pnf.get("control_total"),
            pnf.get("rr") or pnf.get("effect_estimate"),
            pnf.get("ci_lower"), pnf.get("ci_upper"), pnf.get("p_value"),
            pnf.get("source_quote"), pnf.get("source_location")
        ])
        
        # HAT (11 fields - no definition, includes source)
        hat = data.get("hat", {})
        row_data.extend([
            hat.get("reported"),
            hat.get("intervention_events"), hat.get("intervention_total"),
            hat.get("control_events"), hat.get("control_total"),
            hat.get("rr") or hat.get("effect_estimate"),
            hat.get("ci_lower"), hat.get("ci_upper"), hat.get("p_value"),
            hat.get("source_quote"), hat.get("source_location")
        ])
        
        # Retransplantation (12 fields - includes source)
        retx = data.get("retransplantation", {})
        row_data.extend([
            retx.get("reported"), retx.get("timeframe"),
            retx.get("intervention_events"), retx.get("intervention_total"),
            retx.get("control_events"), retx.get("control_total"),
            retx.get("rr") or retx.get("effect_estimate"),
            retx.get("ci_lower"), retx.get("ci_upper"), retx.get("p_value"),
            retx.get("source_quote"), retx.get("source_location")
        ])
        
        # AKI (12 fields - includes source)
        aki = data.get("aki", {})
        row_data.extend([
            aki.get("reported"), aki.get("definition"),
            aki.get("intervention_events"), aki.get("intervention_total"),
            aki.get("control_events"), aki.get("control_total"),
            aki.get("rr") or aki.get("effect_estimate"),
            aki.get("ci_lower"), aki.get("ci_upper"), aki.get("p_value"),
            aki.get("source_quote"), aki.get("source_location")
        ])
        
        # RRT (11 fields - no definition, includes source)
        rrt = data.get("rrt", {})
        row_data.extend([
            rrt.get("reported"),
            rrt.get("intervention_events"), rrt.get("intervention_total"),
            rrt.get("control_events"), rrt.get("control_total"),
            rrt.get("rr") or rrt.get("effect_estimate"),
            rrt.get("ci_lower"), rrt.get("ci_upper"), rrt.get("p_value"),
            rrt.get("source_quote"), rrt.get("source_location")
        ])
        
        # PRS (12 fields - includes source)
        prs = data.get("prs", {})
        row_data.extend([
            prs.get("reported"), prs.get("definition"),
            prs.get("intervention_events"), prs.get("intervention_total"),
            prs.get("control_events"), prs.get("control_total"),
            prs.get("rr") or prs.get("effect_estimate"),
            prs.get("ci_lower"), prs.get("ci_upper"), prs.get("p_value"),
            prs.get("source_quote"), prs.get("source_location")
        ])
        
        self._append_row("outcomes", row_data)
    
    def write_continuous_outcomes(self, data: Dict[str, Any]):
        """Write to Continuous_Outcomes sheet"""
        study_id = data.get("study_id")
        
        # Hospital Stay
        hosp = data.get("hospital_stay_days", {})
        # ICU Stay
        icu = data.get("icu_stay_days", {})
        # Graft Survival
        graft = data.get("graft_survival_1yr", {})
        # Patient Survival
        pat = data.get("patient_survival_1yr", {})
        # Utilization
        util = data.get("utilization_rate", {})
        
        row_data = [
            study_id,
            # Hospital Stay (9 fields including source)
            hosp.get("reported") if isinstance(hosp, dict) else None,
            hosp.get("intervention_mean_or_median") if isinstance(hosp, dict) else None,
            hosp.get("intervention_sd_or_iqr") if isinstance(hosp, dict) else None,
            hosp.get("control_mean_or_median") if isinstance(hosp, dict) else None,
            hosp.get("control_sd_or_iqr") if isinstance(hosp, dict) else None,
            hosp.get("measure") if isinstance(hosp, dict) else None,
            hosp.get("p_value") if isinstance(hosp, dict) else None,
            hosp.get("source_quote") if isinstance(hosp, dict) else None,
            hosp.get("source_location") if isinstance(hosp, dict) else None,
            # ICU Stay (9 fields including source)
            icu.get("reported") if isinstance(icu, dict) else None,
            icu.get("intervention_mean_or_median") if isinstance(icu, dict) else None,
            icu.get("intervention_sd_or_iqr") if isinstance(icu, dict) else None,
            icu.get("control_mean_or_median") if isinstance(icu, dict) else None,
            icu.get("control_sd_or_iqr") if isinstance(icu, dict) else None,
            icu.get("measure") if isinstance(icu, dict) else None,
            icu.get("p_value") if isinstance(icu, dict) else None,
            icu.get("source_quote") if isinstance(icu, dict) else None,
            icu.get("source_location") if isinstance(icu, dict) else None,
            # Graft Survival 1yr (13 fields including source)
            graft.get("reported") if isinstance(graft, dict) else None,
            graft.get("intervention_percent") if isinstance(graft, dict) else None,
            graft.get("control_percent") if isinstance(graft, dict) else None,
            graft.get("intervention_events") if isinstance(graft, dict) else None,
            graft.get("intervention_total") if isinstance(graft, dict) else None,
            graft.get("control_events") if isinstance(graft, dict) else None,
            graft.get("control_total") if isinstance(graft, dict) else None,
            graft.get("hr") if isinstance(graft, dict) else None,
            graft.get("ci_lower") if isinstance(graft, dict) else None,
            graft.get("ci_upper") if isinstance(graft, dict) else None,
            graft.get("p_value") if isinstance(graft, dict) else None,
            graft.get("source_quote") if isinstance(graft, dict) else None,
            graft.get("source_location") if isinstance(graft, dict) else None,
            # Patient Survival 1yr (13 fields including source)
            pat.get("reported") if isinstance(pat, dict) else None,
            pat.get("intervention_percent") if isinstance(pat, dict) else None,
            pat.get("control_percent") if isinstance(pat, dict) else None,
            pat.get("intervention_events") if isinstance(pat, dict) else None,
            pat.get("intervention_total") if isinstance(pat, dict) else None,
            pat.get("control_events") if isinstance(pat, dict) else None,
            pat.get("control_total") if isinstance(pat, dict) else None,
            pat.get("hr") if isinstance(pat, dict) else None,
            pat.get("ci_lower") if isinstance(pat, dict) else None,
            pat.get("ci_upper") if isinstance(pat, dict) else None,
            pat.get("p_value") if isinstance(pat, dict) else None,
            pat.get("source_quote") if isinstance(pat, dict) else None,
            pat.get("source_location") if isinstance(pat, dict) else None,
            # Utilization (7 fields including source)
            util.get("reported") if isinstance(util, dict) else None,
            util.get("intervention_utilized") if isinstance(util, dict) else None,
            util.get("intervention_offered") if isinstance(util, dict) else None,
            util.get("control_utilized") if isinstance(util, dict) else None,
            util.get("control_offered") if isinstance(util, dict) else None,
            util.get("source_quote") if isinstance(util, dict) else None,
            util.get("source_location") if isinstance(util, dict) else None,
            data.get("notes")
        ]
        self._append_row("continuous", row_data)
    
    def write_rob2(self, data: Dict[str, Any]):
        """Write to RoB2_RCT sheet"""
        d1 = data.get("d1_randomization", {})
        d2 = data.get("d2_deviations", {})
        d3 = data.get("d3_missing_data", {})
        d4 = data.get("d4_measurement", {})
        d5 = data.get("d5_selection", {})
        
        def get_quotes(domain_data):
            quotes = domain_data.get("support_quotes", [])
            if isinstance(quotes, list):
                return "; ".join([q.get("quote", q) if isinstance(q, dict) else str(q) for q in quotes])
            return str(quotes) if quotes else None
        
        def get_sq(domain_data, key):
            sq = domain_data.get("signaling_questions", {})
            return sq.get(key) if isinstance(sq, dict) else None
        
        row_data = [
            data.get("study_id"),
            # D1
            d1.get("judgment"),
            get_sq(d1, "random_sequence_generation"),
            get_sq(d1, "allocation_concealment"),
            get_sq(d1, "baseline_imbalances"),
            get_quotes(d1),
            # D2
            d2.get("judgment"),
            get_sq(d2, "participants_blinded"),
            get_sq(d2, "personnel_blinded"),
            get_sq(d2, "intention_to_treat"),
            get_sq(d2, "deviations_occurred"),
            get_quotes(d2),
            # D3
            d3.get("judgment"),
            get_sq(d3, "outcome_data_available"),
            get_sq(d3, "dropout_rate_intervention"),
            get_sq(d3, "dropout_rate_control"),
            get_sq(d3, "dropout_reasons_balanced"),
            get_quotes(d3),
            # D4
            d4.get("judgment"),
            get_sq(d4, "outcome_assessor_blinded"),
            get_sq(d4, "outcome_objective"),
            get_quotes(d4),
            # D5
            d5.get("judgment"),
            get_sq(d5, "preregistered_protocol"),
            data.get("registry_id"),
            get_sq(d5, "outcomes_match_protocol"),
            get_quotes(d5),
            # Overall
            data.get("overall_judgment"),
            data.get("overall_rationale")
        ]
        self._append_row("rob2", row_data)
    
    def write_robins_i(self, data: Dict[str, Any]):
        """Write to ROBINS_I_NRS sheet"""
        d1 = data.get("d1_confounding", {})
        d2 = data.get("d2_selection", {})
        d3 = data.get("d3_classification", {})
        d4 = data.get("d4_deviations", {})
        d5 = data.get("d5_missing_data", {})
        d6 = data.get("d6_measurement", {})
        d7 = data.get("d7_selection", {})
        
        def get_quotes(domain_data):
            quotes = domain_data.get("support_quotes", [])
            if isinstance(quotes, list):
                return "; ".join([q.get("quote", q) if isinstance(q, dict) else str(q) for q in quotes])
            return str(quotes) if quotes else None
        
        def get_sq(domain_data, key):
            sq = domain_data.get("signaling_questions", {})
            return sq.get(key) if isinstance(sq, dict) else None
        
        # Handle confounders list
        confounders = get_sq(d1, "confounders_considered")
        if isinstance(confounders, list):
            confounders = ", ".join(confounders)
        
        row_data = [
            data.get("study_id"),
            # D1
            d1.get("judgment"),
            confounders,
            get_sq(d1, "adjustment_method"),
            get_sq(d1, "residual_confounding_likely"),
            get_quotes(d1),
            # D2
            d2.get("judgment"),
            get_sq(d2, "selection_into_study"),
            get_sq(d2, "exclusions_post_intervention"),
            get_quotes(d2),
            # D3
            d3.get("judgment"),
            get_sq(d3, "classification_based_on"),
            get_sq(d3, "misclassification_possible"),
            get_quotes(d3),
            # D4
            d4.get("judgment"),
            get_sq(d4, "co_interventions_balanced"),
            get_sq(d4, "switches_between_groups"),
            get_quotes(d4),
            # D5
            d5.get("judgment"),
            get_sq(d5, "data_available_for_all"),
            get_sq(d5, "differential_missingness"),
            get_quotes(d5),
            # D6
            d6.get("judgment"),
            get_sq(d6, "outcome_assessors_aware"),
            get_sq(d6, "outcome_ascertainment_comparable"),
            get_quotes(d6),
            # D7
            d7.get("judgment"),
            get_sq(d7, "prespecified_analysis"),
            get_sq(d7, "multiple_analyses_performed"),
            get_quotes(d7),
            # Overall
            data.get("overall_judgment"),
            data.get("overall_rationale")
        ]
        self._append_row("robins", row_data)
    
    def write_notes(self, data: Dict[str, Any], study_id: str, extractor: str = "LLM"):
        """Write to Extraction_Notes sheet"""
        concerns = data.get("data_quality_concerns", [])
        items = data.get("unclear_items_for_review", [])
        overlaps = data.get("potential_overlaps", [])
        
        overlap_study = ""
        overlap_reason = ""
        if overlaps and isinstance(overlaps, list) and len(overlaps) > 0:
            if isinstance(overlaps[0], dict):
                overlap_study = overlaps[0].get("study_id", "")
                overlap_reason = overlaps[0].get("overlap_reason", "")
            else:
                overlap_study = str(overlaps[0])
        
        row_data = [
            study_id,
            datetime.now().strftime("%Y-%m-%d"),
            extractor,
            "; ".join(concerns) if isinstance(concerns, list) else concerns,
            "; ".join(items) if isinstance(items, list) else items,
            overlap_study,
            overlap_reason,
            data.get("eligibility_status"),
            data.get("exclusion_reason"),
            data.get("general_notes")
        ]
        self._append_row("notes", row_data)


# =============================================================================
# Main Extraction Pipeline
# =============================================================================

class DataExtractor:
    def __init__(self, model: str = DEFAULT_MODEL, template_path: str = TEMPLATE_PATH):
        self.llm = LLMClient(model)
        self.template_path = template_path
    
    def read_paper(self, file_path: str) -> str:
        """Read paper content from file (txt or PDF)"""
        path = Path(file_path)
        
        if path.suffix.lower() == '.pdf':
            print(f"Reading PDF: {file_path}")
            text = extract_text_from_pdf(file_path)
            if len(text.strip()) < 500:
                print("Trying OCR fallback...")
                text = extract_text_with_ocr_fallback(file_path)
            return text
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
    
    def quick_screen(self, paper_content: str) -> Dict[str, Any]:
        """Run quick screening"""
        print("Running quick screen...")
        prompt = get_quick_screen_prompt(paper_content)
        return self.llm.extract(prompt, max_tokens=2000)
    
    def full_extraction(self, paper_content: str) -> Dict[str, Any]:
        """Run full data extraction"""
        print("Running full extraction...")
        prompt = get_data_extraction_prompt(paper_content)
        return self.llm.extract(prompt, max_tokens=8000)
    
    def extract_rob(self, paper_content: str, study_type: str) -> Dict[str, Any]:
        """Run RoB-focused extraction"""
        print(f"Extracting RoB information ({study_type})...")
        prompt = get_rob_extraction_prompt(paper_content, study_type)
        return self.llm.extract(prompt, max_tokens=4000)
    
    def process_paper(self, file_path: str, output_path: str, 
                      quick_only: bool = False, include_rob: bool = True,
                      supplementary_files: List[str] = None):
        """Process a single paper and write to Excel
        
        Args:
            file_path: Main paper PDF/text file
            output_path: Output Excel file
            quick_only: Only do quick screening
            include_rob: Include RoB extraction
            supplementary_files: List of supplementary PDF/text files to include
        """
        
        # Read main paper
        paper_content = self.read_paper(file_path)
        print(f"  Read {len(paper_content):,} characters from {Path(file_path).name}")
        
        # Read and append supplementary files
        if supplementary_files:
            for supp_file in supplementary_files:
                try:
                    supp_content = self.read_paper(supp_file)
                    paper_content += f"\n\n{'='*60}\n"
                    paper_content += f"SUPPLEMENTARY MATERIAL: {Path(supp_file).name}\n"
                    paper_content += f"{'='*60}\n\n"
                    paper_content += supp_content
                    print(f"  + Added {len(supp_content):,} characters from {Path(supp_file).name}")
                except Exception as e:
                    print(f"  ⚠️ Warning: Could not read {supp_file}: {e}")
            
            print(f"  Total content: {len(paper_content):,} characters")
        
        # Quick screen first
        print("  [1/3] Quick screening...")
        quick_result = self.quick_screen(paper_content)
        
        study_id = quick_result.get('study_id', 'Unknown')
        design = quick_result.get('design', 'Unknown')
        intervention = quick_result.get('intervention', 'Unknown')
        
        print(f"        → Study: {study_id}")
        print(f"        → Design: {design}")
        print(f"        → Intervention: {intervention}")
        
        if quick_result.get('concerns'):
            print(f"        ⚠️  Concerns: {quick_result.get('concerns')}")
        
        if quick_only:
            return quick_result
        
        # Full extraction
        print("  [2/3] Full data extraction...")
        full_result = self.full_extraction(paper_content)
        
        # Determine study type for RoB
        study_type = "RCT"
        if "study_characteristics" in full_result:
            design = full_result["study_characteristics"].get("study_design", "")
        else:
            design = full_result.get("study_design", full_result.get("design", ""))
        
        if design and "RCT" not in design.upper() and "RANDOM" not in design.upper():
            study_type = "NRS"
        
        # RoB extraction if requested
        rob_result = {}
        if include_rob:
            print(f"  [3/3] RoB extraction ({study_type})...")
            rob_result = self.extract_rob(paper_content, study_type)
        else:
            print("  [3/3] RoB extraction skipped")
        
        # Write to Excel
        print(f"  Writing to Excel...")
        writer = ExcelWriter(self.template_path, output_path)
        
        # Get study_id
        study_id = None
        if "study_characteristics" in full_result:
            study_id = full_result["study_characteristics"].get("study_id")
            writer.write_study_characteristics(full_result["study_characteristics"])
        else:
            study_id = full_result.get("study_id")
            writer.write_study_characteristics(full_result)
        
        # Perfusion settings
        if "perfusion_settings" in full_result:
            perf = full_result["perfusion_settings"]
            perf["study_id"] = study_id
            writer.write_perfusion_settings(perf)
        
        # Time metrics
        if "time_metrics" in full_result:
            time = full_result["time_metrics"]
            time["study_id"] = study_id
            writer.write_time_metrics(time)
        
        # Outcomes
        if "outcome_data" in full_result:
            outcomes = full_result["outcome_data"]
            outcomes["study_id"] = study_id
            writer.write_outcome_data(outcomes)
            writer.write_continuous_outcomes(outcomes)
        elif "outcomes" in full_result:
            outcomes = full_result["outcomes"]
            outcomes["study_id"] = study_id
            writer.write_outcome_data(outcomes)
            writer.write_continuous_outcomes(outcomes)
        
        # RoB
        if rob_result:
            rob_result["study_id"] = study_id
            if study_type == "RCT":
                writer.write_rob2(rob_result)
            else:
                writer.write_robins_i(rob_result)
        
        # Notes
        notes = full_result.get("extraction_notes", {})
        writer.write_notes(notes, study_id)
        
        writer.save()
        
        return {
            "quick_screen": quick_result,
            "full_extraction": full_result,
            "rob_information": rob_result,
            "study_type": study_type
        }
    
    def batch_process(self, input_dir: str, output_path: str):
        """Process all papers in a directory"""
        input_path = Path(input_dir)
        files = list(input_path.glob("*.pdf")) + list(input_path.glob("*.txt"))
        
        print(f"\nFound {len(files)} files to process")
        print(f"Output file: {output_path}")
        print(f"{'='*60}")
        
        results = []
        success_count = 0
        error_count = 0
        
        for i, file in enumerate(files, 1):
            print(f"\n[{i}/{len(files)}] Processing: {file.name}")
            print(f"-" * 40)
            try:
                result = self.process_paper(str(file), output_path, include_rob=True)
                results.append({
                    "file": file.name,
                    "success": True,
                    "study_id": result.get("quick_screen", {}).get("study_id"),
                    "design": result.get("study_type"),
                    "result": result
                })
                success_count += 1
                print(f"  ✓ Success")
            except Exception as e:
                print(f"  ✗ Error: {e}")
                results.append({
                    "file": file.name,
                    "success": False,
                    "error": str(e)
                })
                error_count += 1
        
        # Summary
        print(f"\n{'='*60}")
        print(f"BATCH PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"Total files:    {len(files)}")
        print(f"Successful:     {success_count}")
        print(f"Errors:         {error_count}")
        print(f"Output file:    {output_path}")
        
        if success_count > 0:
            print(f"\nExtracted studies:")
            for r in results:
                if r.get("success"):
                    print(f"  • {r.get('study_id', 'Unknown')} ({r.get('design', '?')}) - {r['file']}")
        
        if error_count > 0:
            print(f"\nFailed files:")
            for r in results:
                if not r.get("success"):
                    print(f"  • {r['file']}: {r.get('error', 'Unknown error')}")
        
        return results


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Data Extractor v2 for Machine Perfusion Systematic Review",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single paper extraction
  python extractor_v2.py -i paper.pdf -o results.xlsx
  
  # With supplementary materials
  python extractor_v2.py -i paper.pdf -s supplementary.pdf -o results.xlsx
  
  # Multiple supplementary files
  python extractor_v2.py -i paper.pdf -s supp1.pdf supp2.pdf -o results.xlsx
  
  # Quick screen only
  python extractor_v2.py -i paper.pdf --quick
  
  # Batch process folder
  python extractor_v2.py -i ./papers/ -o results.xlsx --batch
  
  # Save JSON alongside Excel
  python extractor_v2.py -i paper.pdf -o results.xlsx --json
        """
    )
    parser.add_argument("--input", "-i", required=True, help="Input file or directory")
    parser.add_argument("--supplementary", "-s", nargs="+", help="Supplementary file(s) to include")
    parser.add_argument("--output", "-o", default="extraction_results.xlsx", help="Output Excel file")
    parser.add_argument("--template", "-t", default=TEMPLATE_PATH, help="Template Excel file")
    parser.add_argument("--model", "-m", default=DEFAULT_MODEL, help="Claude model to use")
    parser.add_argument("--quick", action="store_true", help="Quick screen only (no full extraction)")
    parser.add_argument("--no-rob", action="store_true", help="Skip RoB information extraction")
    parser.add_argument("--batch", action="store_true", help="Batch process all files in directory")
    parser.add_argument("--json", action="store_true", help="Also save results as JSON file")
    
    args = parser.parse_args()
    
    print(f"\n{'='*60}")
    print(f"Data Extractor v2 - Machine Perfusion Systematic Review")
    print(f"{'='*60}")
    print(f"Model: {args.model}")
    print(f"Template: {args.template}")
    print(f"Output: {args.output}")
    
    extractor = DataExtractor(model=args.model, template_path=args.template)
    
    if args.batch:
        results = extractor.batch_process(args.input, args.output)
        
        if args.json:
            json_path = args.output.replace('.xlsx', '_batch.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False, default=str)
            print(f"JSON saved to: {json_path}")
    else:
        # Handle supplementary files
        supplementary_files = args.supplementary if args.supplementary else []
        if supplementary_files:
            print(f"Supplementary files: {len(supplementary_files)}")
            for sf in supplementary_files:
                print(f"  • {sf}")
        
        result = extractor.process_paper(
            args.input, 
            args.output,
            quick_only=args.quick,
            include_rob=not args.no_rob,
            supplementary_files=supplementary_files
        )
        
        if args.quick:
            print("\n=== Quick Screen Result ===")
            print(json.dumps(result, indent=2, ensure_ascii=False))
        
        if args.json:
            json_path = args.output.replace('.xlsx', '.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False, default=str)
            print(f"JSON saved to: {json_path}")


if __name__ == "__main__":
    main()
