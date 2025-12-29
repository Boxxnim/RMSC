"""
LLM-based Study Data Extraction Pipeline
For systematic review on ex vivo machine perfusion in ECD liver transplantation

Usage:
    python extractor.py --input paper.pdf --output results.xlsx
    python extractor.py --input papers_folder/ --output results.xlsx --batch
"""

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional
import argparse

# API clients - uncomment the ones you use
# import anthropic
# import google.generativeai as genai
# from openai import OpenAI

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from prompts import (
    SYSTEM_PROMPT,
    EXTRACTION_PROMPT_TEMPLATE,
    VALIDATION_PROMPT_TEMPLATE,
    OVERLAP_CHECK_PROMPT_TEMPLATE
)
from schemas import FULL_EXTRACTION_SCHEMA


# =============================================================================
# Configuration
# =============================================================================

class Config:
    # API Settings
    EXTRACTION_MODEL = "claude-sonnet-4-20250514"  # Primary extraction
    VALIDATION_MODEL = "gemini-2.5-pro"             # Validation/verification
    
    # Anthropic
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
    
    # Google
    GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
    
    # Paths
    TEMPLATE_PATH = "cohort_tracking_template_v3.xlsx"
    OUTPUT_DIR = "extraction_outputs"
    
    # Processing
    MAX_RETRIES = 3
    REQUIRE_DUAL_VERIFICATION = True
    CONFIDENCE_THRESHOLD = "Medium"  # Minimum confidence for auto-accept
    
    # Sheets in template
    REGISTRY_SHEET = "Study Registry"
    ELIGIBILITY_SHEET = "Eligibility Log"
    LINKAGE_SHEET = "Cohort Linkage"


# =============================================================================
# API Wrappers
# =============================================================================

def call_claude(prompt: str, system: str = SYSTEM_PROMPT) -> str:
    """Call Anthropic Claude API"""
    import anthropic
    
    client = anthropic.Anthropic(api_key=Config.ANTHROPIC_API_KEY)
    
    response = client.messages.create(
        model=Config.EXTRACTION_MODEL,
        max_tokens=4096,
        system=system,
        messages=[{"role": "user", "content": prompt}]
    )
    
    return response.content[0].text


def call_gemini(prompt: str, system: str = SYSTEM_PROMPT) -> str:
    """Call Google Gemini API"""
    import google.generativeai as genai
    
    genai.configure(api_key=Config.GOOGLE_API_KEY)
    model = genai.GenerativeModel(
        model_name=Config.VALIDATION_MODEL,
        system_instruction=system
    )
    
    response = model.generate_content(prompt)
    return response.text


def extract_json_from_response(response: str) -> dict:
    """Extract JSON from LLM response, handling markdown code blocks"""
    # Try to find JSON in code blocks
    json_match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', response)
    if json_match:
        json_str = json_match.group(1)
    else:
        # Try to find raw JSON
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            json_str = json_match.group(0)
        else:
            raise ValueError("No JSON found in response")
    
    return json.loads(json_str)


# =============================================================================
# Core Extraction Logic
# =============================================================================

def extract_study_data(paper_content: str, paper_id: str = "") -> dict:
    """
    Extract structured data from a paper using LLM
    
    Args:
        paper_content: Text content of the paper
        paper_id: Optional identifier for logging
    
    Returns:
        Extracted data dictionary
    """
    prompt = EXTRACTION_PROMPT_TEMPLATE.format(paper_content=paper_content)
    
    for attempt in range(Config.MAX_RETRIES):
        try:
            response = call_claude(prompt)
            data = extract_json_from_response(response)
            
            # Validate required fields
            if "registry_data" not in data or "eligibility" not in data:
                raise ValueError("Missing required sections in extraction")
            
            data["_metadata"] = {
                "extracted_at": datetime.now().isoformat(),
                "extraction_model": Config.EXTRACTION_MODEL,
                "paper_id": paper_id,
                "attempt": attempt + 1
            }
            
            return data
            
        except Exception as e:
            print(f"Extraction attempt {attempt + 1} failed: {e}")
            if attempt == Config.MAX_RETRIES - 1:
                raise
    
    return {}


def validate_extraction(paper_summary: str, extracted_data: dict) -> dict:
    """
    Validate extraction using second model
    
    Args:
        paper_summary: Brief summary of paper for context
        extracted_data: Previously extracted data
    
    Returns:
        Validation results
    """
    prompt = VALIDATION_PROMPT_TEMPLATE.format(
        paper_summary=paper_summary,
        extracted_json=json.dumps(extracted_data, indent=2)
    )
    
    try:
        response = call_gemini(prompt)
        validation = extract_json_from_response(response)
        
        validation["_metadata"] = {
            "validated_at": datetime.now().isoformat(),
            "validation_model": Config.VALIDATION_MODEL
        }
        
        return validation
        
    except Exception as e:
        print(f"Validation failed: {e}")
        return {
            "validation_status": "Error",
            "error": str(e),
            "final_recommendation": "Manual review required"
        }


def check_overlap(study_a: dict, study_b: dict) -> dict:
    """
    Check for potential patient overlap between two studies
    """
    prompt = OVERLAP_CHECK_PROMPT_TEMPLATE.format(
        study_a_summary=json.dumps(study_a, indent=2),
        study_b_summary=json.dumps(study_b, indent=2)
    )
    
    try:
        response = call_claude(prompt)
        return extract_json_from_response(response)
    except Exception as e:
        print(f"Overlap check failed: {e}")
        return {"overlap_type": "Unknown", "error": str(e)}


# =============================================================================
# Excel Integration
# =============================================================================

class ExcelWriter:
    """Handles writing extracted data to Excel template"""
    
    def __init__(self, template_path: str, output_path: str):
        if os.path.exists(template_path):
            self.wb = load_workbook(template_path)
        else:
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        self.output_path = output_path
        self._setup_styles()
    
    def _setup_styles(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.pending_fill = PatternFill('solid', fgColor='FFEB9C')
        self.include_fill = PatternFill('solid', fgColor='C6EFCE')
        self.exclude_fill = PatternFill('solid', fgColor='FFC7CE')
    
    def _get_next_row(self, sheet_name: str) -> int:
        """Find next empty row in sheet"""
        ws = self.wb[sheet_name]
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=1).value is None:
                return row
        return ws.max_row + 1
    
    def write_registry(self, data: dict) -> int:
        """Write to Study Registry sheet"""
        ws = self.wb[Config.REGISTRY_SHEET]
        row = self._get_next_row(Config.REGISTRY_SHEET)
        
        reg = data.get("registry_data", {})
        
        # Map fields to columns (adjust based on your template)
        fields = [
            reg.get("study_id", ""),
            reg.get("first_author", ""),
            reg.get("year", ""),
            reg.get("title", ""),
            reg.get("journal", ""),
            reg.get("doi", ""),
            reg.get("study_design", ""),
            ", ".join(reg.get("centers", [])),
            ", ".join(reg.get("countries", [])),
            reg.get("enrollment_period", ""),
            reg.get("registry_id", ""),
            "",  # Parent Study ID - to be filled manually
            "",  # Relationship Type - to be filled manually
            reg.get("sample_size", {}).get("total", ""),
            reg.get("sample_size", {}).get("intervention", ""),
            reg.get("sample_size", {}).get("control", ""),
            reg.get("intervention_type", ""),
            reg.get("notes", "")
        ]
        
        for col, value in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.thin_border
        
        return row
    
    def write_eligibility(self, data: dict, reviewer: str = "LLM") -> int:
        """Write to Eligibility Log sheet"""
        ws = self.wb[Config.ELIGIBILITY_SHEET]
        row = self._get_next_row(Config.ELIGIBILITY_SHEET)
        
        reg = data.get("registry_data", {})
        elig = data.get("eligibility", {})
        
        fields = [
            reg.get("study_id", ""),
            reg.get("first_author", ""),
            reg.get("year", ""),
            elig.get("decision", "Pending"),
            elig.get("exclusion_category", ""),
            elig.get("rationale", ""),
            elig.get("pico_violation", ""),
            elig.get("precedent", ""),
            reviewer,
            datetime.now().strftime("%Y-%m-%d"),
            "No"  # Reviewer Confirmed
        ]
        
        for col, value in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.thin_border
            
            # Color code by decision
            if col == 4:  # Decision column
                if value == "Include":
                    cell.fill = self.include_fill
                elif value == "Exclude":
                    cell.fill = self.exclude_fill
                else:
                    cell.fill = self.pending_fill
        
        return row
    
    def write_linkage(self, overlap_data: dict) -> int:
        """Write to Cohort Linkage sheet"""
        ws = self.wb[Config.LINKAGE_SHEET]
        row = self._get_next_row(Config.LINKAGE_SHEET)
        
        fields = [
            overlap_data.get("study_id_a", ""),
            overlap_data.get("study_id_b", ""),
            overlap_data.get("overlap_type", ""),
            overlap_data.get("overlap_evidence", ""),
            ", ".join(overlap_data.get("shared_elements", {}).get("centers", [])),
            overlap_data.get("shared_elements", {}).get("enrollment_overlap", ""),
            overlap_data.get("estimated_overlap_percent", ""),
            overlap_data.get("resolution_recommendation", ""),
            "",  # Selected Study - manual
            overlap_data.get("resolution_rationale", ""),
            ""   # Verified By - manual
        ]
        
        for col, value in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.thin_border
        
        return row
    
    def save(self):
        """Save workbook"""
        self.wb.save(self.output_path)
        print(f"Saved to: {self.output_path}")


# =============================================================================
# Pipeline Orchestration
# =============================================================================

def process_paper(
    paper_content: str,
    paper_id: str,
    excel_writer: ExcelWriter,
    validate: bool = True
) -> dict:
    """
    Full extraction pipeline for a single paper
    
    Returns:
        Complete results including extraction, validation, and row numbers
    """
    results = {
        "paper_id": paper_id,
        "status": "processing",
        "extraction": None,
        "validation": None,
        "rows_written": {}
    }
    
    try:
        # Step 1: Extract
        print(f"Extracting: {paper_id}")
        extraction = extract_study_data(paper_content, paper_id)
        results["extraction"] = extraction
        
        # Step 2: Validate (optional)
        if validate and Config.REQUIRE_DUAL_VERIFICATION:
            print(f"Validating: {paper_id}")
            # Create brief summary for validation
            reg = extraction.get("registry_data", {})
            summary = f"{reg.get('first_author')} {reg.get('year')}: {reg.get('title', '')[:100]}"
            
            validation = validate_extraction(summary, extraction)
            results["validation"] = validation
            
            # Check if revision needed
            if validation.get("final_recommendation") == "Revise extraction":
                print(f"Warning: Validation suggests revision for {paper_id}")
                extraction["_needs_review"] = True
        
        # Step 3: Write to Excel
        reg_row = excel_writer.write_registry(extraction)
        elig_row = excel_writer.write_eligibility(extraction)
        
        results["rows_written"] = {
            "registry": reg_row,
            "eligibility": elig_row
        }
        
        results["status"] = "complete"
        
    except Exception as e:
        results["status"] = "error"
        results["error"] = str(e)
        print(f"Error processing {paper_id}: {e}")
    
    return results


def run_batch_extraction(
    input_dir: str,
    output_path: str,
    template_path: str = Config.TEMPLATE_PATH
):
    """
    Process multiple papers from a directory
    """
    input_path = Path(input_dir)
    papers = list(input_path.glob("*.txt")) + list(input_path.glob("*.pdf"))
    
    print(f"Found {len(papers)} papers to process")
    
    writer = ExcelWriter(template_path, output_path)
    all_results = []
    
    for paper_path in papers:
        paper_id = paper_path.stem
        
        # Read content (for PDFs, you'd need PyMuPDF or similar)
        if paper_path.suffix == ".txt":
            with open(paper_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            # TODO: Add PDF extraction
            print(f"Skipping PDF (add PDF extraction): {paper_path}")
            continue
        
        results = process_paper(content, paper_id, writer)
        all_results.append(results)
    
    writer.save()
    
    # Summary
    success = sum(1 for r in all_results if r["status"] == "complete")
    errors = sum(1 for r in all_results if r["status"] == "error")
    print(f"\nBatch complete: {success} successful, {errors} errors")
    
    return all_results


# =============================================================================
# CLI Interface
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Extract study data for systematic review"
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Input file or directory"
    )
    parser.add_argument(
        "--output", "-o",
        default="extraction_results.xlsx",
        help="Output Excel file"
    )
    parser.add_argument(
        "--template", "-t",
        default=Config.TEMPLATE_PATH,
        help="Excel template path"
    )
    parser.add_argument(
        "--batch", "-b",
        action="store_true",
        help="Process directory of files"
    )
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip validation step"
    )
    
    args = parser.parse_args()
    
    if args.batch:
        run_batch_extraction(
            args.input,
            args.output,
            args.template
        )
    else:
        # Single file
        with open(args.input, 'r', encoding='utf-8') as f:
            content = f.read()
        
        writer = ExcelWriter(args.template, args.output)
        results = process_paper(
            content,
            Path(args.input).stem,
            writer,
            validate=not args.no_validate
        )
        writer.save()
        
        print(json.dumps(results, indent=2, default=str))


if __name__ == "__main__":
    main()
